import openpyxl, json, re, sys
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

def parse_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    cups = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        position_cells = []
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if val == 'Position':
                    position_cells.append((ri, ci))
        for pos_row, pos_col in position_cells:
            cup_name = None
            for sr in range(pos_row - 1, max(pos_row - 5, -1), -1):
                val = rows[sr][pos_col] if pos_col < len(rows[sr]) else None
                if val and (str(val).startswith('COTD') or str(val).startswith('COTW')):
                    cup_name = str(val).strip()
                    break
            if not cup_name: continue
            players = []
            last_pos = None
            for row in rows[pos_row + 1:]:
                if pos_col >= len(row) or pos_col + 1 >= len(row): continue
                pos, name = row[pos_col], row[pos_col + 1]
                if name is None: continue
                name_str = str(name).strip()
                if name_str.startswith('*'):  # Skip Lexer's spreadsheet notes
                    continue
                if pos is not None:
                    try:
                        pos_clean = str(pos).rstrip('*').strip()
                        last_pos = int(float(pos_clean))
                        players.append((last_pos, name_str))
                    except: continue
                elif last_pos is not None:
                    # No position number = tied with previous position
                    players.append((last_pos, name_str))
            if players:
                cups.append({'name': cup_name, 'players': sorted(players, key=lambda x: x[0])})
    return cups

def parse_troll_cups(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Troll COTDs']
    rows = list(ws.iter_rows(values_only=True))
    troll_defs = [
        ('Troll COTD 1', 15.5, 0), ('Troll COTD 2', 26.5, 6),
        ('Troll COTD 3', 36.5, 12), ('Troll COTD 4', 41.5, 17),
        ('Troll COTD 5', 44.5, 23), ('Troll COTD 6', 48.5, 29),
        ('Troll COTD 7', 50.5, 35), ('Troll COTD 8', 56.5, 41),
        ('Troll COTW 9', 63.5, 47), ('Troll COTD 10', 71.5, 53),
        ('Troll COTD 11', 88.5, 59),
    ]
    cups = []
    # Override troll cup 4 with reconstructed lobby from screenshots
    troll4_override = [
        (1, 'justMaki'), (2, 'Mark'), (3, 'Lexer'),
        (3, '[CTR]R0nanC'), (3, 'ButItRuns'), (3, 'St Nicholas'), (3, 'Hi Im Yolo'),
        (3, 'RoundNzt'), (3, '[CTR]L3it3R'), (3, 'FINSTER83'), (3, 'Neb'),
        (3, '[CTR]Hydro'), (3, '[RTR]Fwogiie'), (3, '[heyo]Mr. Hubub'), (3, 'agix'),
        (3, 'koz'), (3, '[6dog] schmxrg'), (3, 'Lynhardt'),
    ]
    for name, order, pc in troll_defs:
        players = []
        for row in rows[5:]:
            if pc >= len(row) or pc+1 >= len(row): continue
            pos, nm = row[pc], row[pc+1]
            if pos is None or nm is None: continue
            nm = str(nm).strip()
            if 'Other Player' in nm: continue
            # Fix malformed tags
            nm = nm.replace('[CTR[', '[CTR]').replace('[RTR[', '[RTR]')
            try: players.append((int(float(pos)), nm))
            except: continue
        if len(players) >= 2:
            cups.append({'name': name, 'players': sorted(players, key=lambda x: x[0])})
        # Override troll cup 4 with screenshot-reconstructed lobby
        if name == 'Troll COTD 4':
            cups[-1 if cups and cups[-1]['name'] == 'Troll COTD 4' else len(cups):] = []
            cups.append({'name': name, 'players': troll4_override})
    return cups

import os
_dir = os.path.dirname(os.path.abspath(__file__))
_p = lambda f: os.path.join(_dir, f)
all_cups = (parse_file(_p('Zeepkist COTDs 1-25.xlsx')) +
            parse_file(_p('Zeepkist COTDs 26-50.xlsx')) +
            parse_file(_p('Zeepkist COTDs 51-75.xlsx')) +
            parse_file(_p('COTDs 76-100.xlsx')) +
            parse_file(_p('COTDs 101-125.xlsx')) +
            parse_file(_p('COTD 126-130.xlsx')) +
            parse_file(_p('COTD 131-138.xlsx')) +
            parse_file(_p('cup roulette.xlsx')) +
            parse_troll_cups(_p('Troll cup.xlsx')))

# Extract cup number for sorting - handle both COTD and COTW and specials
SPECIAL_CUP_ORDER = {
    'COTD Roulette 1': 25.5,
    'COTD Roulette 2': 65.5,
    'Troll COTD 1': 15.5,
    'Troll COTD 2': 26.5,
    'Troll COTD 3': 36.5,
    'Troll COTD 4': 41.5,
    'Troll COTD 5': 44.5,
    'Troll COTD 6': 48.5,
    'Troll COTD 7': 50.5,
    'Troll COTD 8': 56.5,
    'Troll COTW 9': 63.5,
    'Troll COTD 10': 71.5,
    'Troll COTD 11': 88.5,
}
def cup_num(name):
    if name in SPECIAL_CUP_ORDER:
        return SPECIAL_CUP_ORDER[name]
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else 0
all_cups.sort(key=lambda c: cup_num(c['name']))

# Deduplicate (same cup number)
seen = set()
deduped = []
for c in all_cups:
    n = cup_num(c['name'])
    if n not in seen:
        seen.add(n)
        deduped.append(c)
all_cups = deduped

# Filter non-standard cups (Troll COTDs, Roulette)
def is_nonstandard(cup_name):
    return cup_name.startswith('Troll ') or 'Roulette' in cup_name

pure_cups = [c for c in all_cups if not is_nonstandard(c['name'])]
print(f"Pure cups: {len(pure_cups)} (excluded {len(all_cups) - len(pure_cups)} non-standard)")

# Fix COTD 16: both jandje and justMaki DNF'd the final - tied at 2nd, no 1st place
for c in all_cups:
    if c['name'] == 'COTD 16':
        c['players'] = [(2 if pos == 1 else pos, name) for pos, name in c['players']]
        break

print(f"Parsed {len(all_cups)} cups")
for c in all_cups:
    print(f"  {c['name']}: {len(c['players'])} players")

# Collect all names for alias detection
all_names = set()
for cup in all_cups:
    for _, name in cup['players']:
        all_names.add(name)

def strip_tag(name):
    return re.sub(r'\[.*?\]\s*', '', name).strip()

# Build comprehensive alias map
CANONICAL = {
    '376': ['376.0'],
    'AndMe': ['[ORIG]AndMe16', 'AndMe14', 'AndMe15', 'AndMe16', 'AndMe17', '[COMY]AndMe17', '[CSC]AndMe17'],
    'bernhard': ['[Lord] bernhard'],
    'Butter': ['[ZST] Butter'],
    'Codewalt': ['CodeWalt'],
    'DragonBoi': ['[Top3]DragonBoi'],
    'Naomi': ['[POIN]Fwogiie','[RTR]Fwogiie','[Tran]Fwogiie','[frog]Fwogiie','[RTR]Fwogiie.Kawaii','Fwogiie.Kawaii','Fwogiie','[ASJE]Naomi'],
    'GuAlexItar': ['[RTR]GuAlexItar'],
    'Hi Im Yolo': ['[HRR]Hi Im Yolo','[RIP]Hi Im Yolo'],
    'Hydro': ['[BFP] Hydro','[CTR]Hydro','[RTR] Hydro','[SLOW] Hydro','k2blue','[WOW]Hydro','[ZOMN] Hydro','[CTR[Hydro','MystiCookies','MysticVoid','l3purple'],
    'I_stay_sideway': ['[RTR] I_stay_sideway','Istaysideway','[RTR]Istaysideway'],
    'ITz_WillleeMan': ['ITz_Willleeman','Itz_WillleeMan'],
    'jandje': ['[BFP] jandje','[CTR] jandje','[CTR]jandje'],
    'Joking': ['[BGR] Joking'],
    'justMaki': ['[test] justMaki'],
    'Kernkob': ['[CTR]Kernkob'],
    'L3it3R': ['L3it3r','[CTR] L3it3R','[CTR]L3it3R','[CTR]L3it3r'],
    'Last': ['[dumb] Last'],
    'Lazy_Echidna': ['[NIL]Lazy_Echidna','[TOG]Lazy_Echidna','[TOG]Lazy_echidna'],
    'Mark': ['[RTR]Mark'],
    'Metalted': ['[ZMS] Metalted','Matalted'],
    'Not That Guy': ['[GANS] Not That Guy','[RTR] Not That Guy'],
    'OwlPlague': ['[CTR] OwlPlague','[CTR]OwlPlague'],
    'Pants': ['[COLD]Pants'],
    'Phoenjx': ['[HUGS]Phoenjx'],
    'Pigbuy': ['[OREO] Pigbuy','[OREO]Pigbuy','[OR]Pigbuy'],
    'Principe': ['[GV] Principe'],
    'Quickracer10': ['[KURK] Quickracer10','[AJSE] Quickracer10','quickracer10'],
    'R0nanC': ['[CTR] R0nanC','[CTR]R0nanC','R0nanc'],
    'readfreak7': ['[HRR] readfreak7','[HRR]readfreak7','[PFE] readfreak7'],
    'Renergy': ['[TOOB] Renergy','[TOOB]Renergy','[UP] Renergy','[WSHD] Renergy','[just] Renergy','[meh] Renergy'],
    'rsgold': ['[OR] rsgold'],
    'rtyyyyb': ['[TOB] rtyyyyb','[ZET] rtyyyyb','[ZET]rtyyyyb','[TBD] rtyyyyb','rtube','[SUCK] rtube','[TBD] rtube','[TBD]rtube','[BAP]rtube','[Toob]rtube','[dumb]rtube','[sad]rtube'],
    'Sandals': ['[CTR] Sandals'],
    'schmxrg': ['[6dog] schmxrg','[dogg] schmxrg','[goat] schmxrg'],
    'shadeely': ['Shadeely','[KBR] shadeely','[RonC] shadeely'],
    'St Nicholas': ['St NIcholas','[cozy]St Nicholas'],
    'TheBamboozler': ['[CTR] TheBamboozler','[CTR]TheBamboozler'],
    'TheBestMaidens': ['[TBR]TheBestMaidens','[WAM]TheBestMaidens','[WAM] TheBestMaidens'],
    'TraNin': ['[P50] TraNin','[TD38] TraNin'],
    'Tritonas1237': ['[zfwo]Tritonas1237'],
    'TwoFace': ['[CTR]TwoFace','[CTR]Twoface'],
    'Warsnac': ['[CHR]Warcans', 'Warcans', '[BAP]Warsnac', '[CHR]Warsnac', '[old]Warsnac'],
    'zivecef': ['[WAM] zivecef','[ZST] zivecef','[ZST]zivecef'],
    'ZOMAN': ['[ARMS] ZOMAN','[Bath] ZOMAN','[Blub] ZOMAN','[Choo] ZOMAN',
              '[DNF] Did Not DNF ZOMAN','DNF artist (ZOMAN)',
              '[KUNG] ZOMAN','[Kung] ZOMAN','[SLOW]ZOMAN','[SNTA] ZOMAN',
              '[TOOB] ZOMAN','[oOOo] ZOMAN'],
    'Zodiak': ['im washed [zodiakism]'],
    'Cbad Cruiser': ['[CTR]Cbad Cruiser'],
    'LupensCruor': ['[CTR] LupensCruor'],
    'NathWalt': ['[CTR] NathWalt', '[CRT]NathWalt', '[CTR]NathWalt'],
    '3rdseyeview': ['[vibe]3rdseyeview'],
    'A Broken Forklift': ['[FORK] A Broken Forklift'],
    'a_random_tumbleweed': ['[ZET] a_random_tumbleweed'],
    'Axo': ['[RDX] axo'],
    'Hidef09': ['[badR] Hidef09', '[badr] Hidef09'],
    'icRS': ['[LATE] icRS'],
    'LILWOOLEY': ['[ZET]LILWOOLEY', '[ZST]LILWOOLEY'],
    'Lynhardt': ['[KBW] Lynhardt'],
    'Odist': ['[WAM]Odist'],
    'PandaMane': ['[FOV]PandaMane', '[FPV]PandaMane'],
    'PoopSheriff': ['[CTR] PoopSheriff', '[CTR]PoopSheriff'],
    'Psycho No. 7': ['[PTSD]Psycho No. 7'],
    'Quickracer10': ['[KURK] Quickracer10','[AJSE] Quickracer10','quickracer10','[ASJE] Quickracer10','[CC] Quickracer10'],
    'Roader': ['[BOB]Roader', '[OR] Roader'],
    'Shadynook': ['[LATE]Shadynook'],
    'Stick': ['[ZET]Stick'],
    'Striking Vyper': ['[CTR]Striking Vyper'],
    'XpERt': ['[TBD]XpERt'],
    'Mμ': ['[CTR]Mμ', 'Mu'],
    'Kaiser64': ['[TEA]Kaiser64'],
    'LKat': ['[GGG]Lkat','[GGG]LKat','[MMM]LKat'],
    'void': ['[ZET]void','[poop]void','[bob]void','[popo]void'],
    'Achmetha0626': ['[ERR]Achmetha0626'],
    'Ax1ss': ['[FaS] Ax1ss','[TOG] Ax1ss'],
    'BootyMcShooty88': ['[SBOI] BootyMcShooty88'],
    'Chinpokomon': ['[RB] Chinpokomon','[RB]Chinpokomon'],
    'CopperFeather': ['[WHI] CopperFeather'],
    'DBNULL': ['[ZAGA]DBNULL'],
    'DocRee': ['[CLWN]DocRee'],
    'dudeeitsraymond': ['[iDad] dudeeitsraymond','[iDad]dudeeitsraymond'],
    'Exterminate': ['[XTR] Exterminate','[XTR]Exterminate'],
    'FlyBoy': ['Fly8oy'],
    'Gimpel': ['Leviathan (Gimpel)'],
    'GuillaumePN': ['[Qc] GuillaumePN'],
    'incredulouspotato': ['[PINK]incredulouspotato','[TTV]incredulouspotato'],
    'ioi8': ['[TOG][KBR] ioi8','[TOG]ioi8'],
    'it_is_nic': ['[its] it_is_nic'],
    'Jeffrey': ['[BOGO]Jeffrey'],
    'Meowbee': ['[HRT!]Meowbee'],
    'Pheonjx': ['[HUGS]Pheonjx'],
    'Pilaf': ['[PILF] Pilaf'],
    'Sheriff': ['[Poop]Sheriff'],
    'Smullie': ['[KURK] Smullie','[KURK]Smullie'],
    'ttv/Lilly the Bun': ['[bnuy]ttv/Lilly the Bun'],
    'Ulv_RaVn': ['[XTR] [VK] Ulv_RaVn','[XTR] [VK] Ulv_Ravn'],
    'vectortrajector': ['[ZET]vectortrajector'],
    'Victor': ['[GGG]Victor','[MMM]Victor'],
    "Zeke Ryu'kai": ["[BoF3] Zeke Ryu'kai"],
    '=XDC=WOLF': ['=XDC=Wolf'],
    'tws20so': ['[DRFT] tws20so', '[IRS] tws20so'],
    'IronDragon111000': ['[CSC]IronDragon111000'],
    'Mokster': ['[CSC]Mokster'],
    'vortex': ['[mib]vortex'],
    'Lexer': ['[BRIT] Lexer'],
    'Murrl': ['[BAP]Murrl', '[Burp]Murrl', '[Toob]Murrl', 'MeroMeroNoMi', '[Mero] Murri'],
    'JakeAdjacent': ['[CD]SadD0ge', '[SWMG]SadD0ge', '[SWMG]JakeAdjacent', 'SadD0ge'],
    'Beans': ['[CTR]Beans'],
    'DeeDeeNaNaNa': ['[CSC] DeeDeeNaNaNa', '[CSC]DeeDeeNaNaNa'],
    'Form': ['[fn]Form'],
    'Hellmet': ['[Dark]Hellmet', '[ZOFC]Hellmet'],
    'Jakie': ['[CD] Jakie', '[ZET] Jakie'],
    'K410K3N': ['[20X]K410K3N', '[20x]K410K3N', '[Gwen]K410K3N'],
    'loganbradley714': ['[GFHL]loganbradley714', '[GLHF]loganbradley714'],
    'lucanakin': ['[DNFF]lucanakin'],
    'MarcSubstitute': ['[DHLU]MarcSubstitute', '[SLOW]MarcSubstitute'],
    'MetalCJ': ['[TTR]MetalCJ'],
    'microways': ['[Quac] microways'],
    'MMXD18': ['[Toob]MMXD18'],
    'Moody': ['[CTR]Moody', '[MIB]Moody'],
    'RadAbsRad': ['[Meow]RadAbsRad'],
    'Redstony': ['[Stc3]Redstony', '[TILT]Redstony'],
    'Six': ['SixSixSevenSeven', '[BAP]SixSixSevenSeven', '[BAP]Six'],
    'Sterben': ['[BAP]Sterben','[PNCK]Mini P.E.K.K.A','λ','Lλmbda','[FPV]Lλmbda','[PCDJ]Sterben'],
    'stindt': ['[KAAS]stindt', '[Lame]stindt', '[Lame}stindt', '[Slow]stindt', '[Tame]stindt'],
    'Tommygaming': ['[CSC]Tommygaming', '[OOPS]Tommygaming', '[jofk]Tommygaming', 'TommyGaming5132', 'Tommygaming5132', '[TG]Tommygaming5132', '[CSC]Tommygaming5132', '[CSC]Tommygaming6132', '[C3PO]Tommygaming5132'],
    'WotterBytes': ['Wotterbytes'],
    'ping': ['[bad] ping', '[boom] ping', '[no] ping', '[pong]ping'],
    '//////void': ['[gorp]//////void'],
    'agix': ['[GYMC] agix'],
    'An Actual g00se': ['[CSC] An Actual g00se', '[Err] An Actual g00se', '[CSC] BaBa is g00se', '[CSC] CantFindTheg00se'],
    'BOB THE GAMER': ['[BOGO]BOB THE GAMER', '[MEAT]BOB THE GAMER'],
    'ferinine': ['[Err]ferinine'],
    'frenchteost': ['[LEXR]frenchteost', '[ZET]frenchteost'],
    'Ionjig': ['ionjig'],
    'Lamp': ['[CTR]Lamp', '[The]Lamp', '[bam]Lamp'],

    'MackCheesy': ['[CHEZ]MackCheesy', '[ZET]MackCheesy'],
    'magostinho20': ['[F1] magostinho20', '[I290] magostinho20'],
    'MeOne2Three4': ['meone2three4'],
    'Mortishade': ['[CTR]Mortishade', '[bam]Mortishade'],
    'Mr. Hubub': ['[Heyo]Mr. Hubub', '[heyo]Mr. Hubub'],
    'OLR94': ['[CSC]OLR94'],
    'Ploddip': ['[ZET] Ploddip', '[ZET]Ploddip'],
    'R0nanC': ['R0nanc'],
    'redal': ['[CSC] redal'],
    'SkyVirus': ['[NOOB]SkyVirus'],
    'Socks242': ['[Fly] Socks242'],
    'variableferret': ['[CSC] variableferret'],
    'Weak_Knees': ['[COMY]Weak_Knees','Weak_knees'],
    'Wheelie': ['[ZET] Wheelie', '[ZET]Wheelie'],
    'Zachafinackus': ['[Sumo]Zachafinackus'],
    'Heart-TGV': ['[TTR]Heart-TGV'],
}

NAME_MAP = {}
for canonical, aliases in CANONICAL.items():
    for alias in aliases:
        NAME_MAP[alias] = canonical

# Auto-extend: detect any new [TAG]Name patterns where Name already exists as canonical
for n in all_names:
    stripped = strip_tag(n)
    if stripped != n and stripped in all_names and n not in NAME_MAP:
        # Check if stripped version is a canonical
        is_canonical = stripped in CANONICAL or any(stripped == c for c in CANONICAL)
        if is_canonical:
            NAME_MAP[n] = stripped

def normalize(name):
    return NAME_MAP.get(name, name)

for cup in all_cups:
    cup['players'] = [(pos, normalize(name)) for pos, name in cup['players']]

# Handle shared-account entries: "account_name (elo=RealPlayer)"
# Spreadsheet tag format: rtm_lover2007 (elo=Kernkob)
# RealPlayer gets ELO credit; account_name becomes a ghost (history only, no ELO effect)
import re as _re
for cup in all_cups:
    new_players = []
    ghosts = []
    for pos, name in cup['players']:
        m = _re.match(r'^(.+?)\s*\(elo=(.+?)\)$', name)
        if m:
            ghost_name = m.group(1).strip()
            real_name = normalize(m.group(2).strip())
            new_players.append((pos, real_name))
            ghosts.append((pos, ghost_name, real_name))
            print(f"  Ghost split in {cup['name']}: {ghost_name} → ELO:{real_name}, ghost:{ghost_name}")
        else:
            new_players.append((pos, name))
    cup['players'] = new_players
    cup['ghosts'] = ghosts  # stored separately for history-only processing

# Check remaining duplicates
by_stripped = defaultdict(set)
for cup in all_cups:
    for _, name in cup['players']:
        by_stripped[strip_tag(name).lower()].add(name)
remaining = {k: v for k, v in by_stripped.items() if len(v) > 1}
if remaining:
    print(f"\nWARNING: {len(remaining)} unresolved duplicates:")
    for k, v in sorted(remaining.items()):
        print(f"  {k}: {sorted(v)}")
else:
    print("\nAll duplicates resolved!")

# Count unique
unique = set()
for cup in all_cups:
    for _, name in cup['players']:
        unique.add(name)
print(f"Unique players: {len(unique)}")

# === ELO ===
STARTING = 1500; K_BASE = 32; PROV_CUPS = 12; PROV_MULT = 1.5
DECAY = 0.995; GRACE = 3

def E(ra,rb): return 1.0/(1.0+10.0**((rb-ra)/400.0))

def pct_mult(pos, n):
    pct = pos / n
    if pct <= 0.08: return 3.0
    if pct <= 0.15: return 2.0
    if pct <= 0.25: return 1.3
    if pct <= 0.50: return 0.8
    return 0.5

def compute_standard_elo(cups, no_ghosts=False):
    ratings = defaultdict(lambda: STARTING)
    gp = defaultdict(int); history = defaultdict(list)
    wins = defaultdict(int); pods = defaultdict(lambda:[0,0,0])
    best = defaultdict(lambda:999); total_pos = defaultdict(int); avg_cups = defaultdict(int)
    for cup in cups:
        players = cup['players']; n = len(players)
        if n < 2: continue
        deltas = defaultdict(float)
        for i in range(n):
            pi,ni = players[i]; ra = ratings[ni]
            k = K_BASE/(n-1)
            if gp[ni] < PROV_CUPS: k *= PROV_MULT
            for j in range(n):
                if i==j: continue
                pj,nj = players[j]
                e = E(ra, ratings[nj])
                s = 1.0 if pi<pj else (0.0 if pi>pj else 0.5)
                deltas[ni] += k*(s-e)
        for pos,name in players:
            ratings[name] += deltas[name]; gp[name] += 1
            history[name].append({'cup':cup['name'],'position':pos,'rating':round(ratings[name],1),'lobby_size':n})
            is_troll4_dnf = cup['name'] == 'Troll COTD 4' and pos == 3
            if not is_troll4_dnf:
                total_pos[name] += pos
                avg_cups[name] += 1
                if pos < best[name]: best[name] = pos
            if pos==1: wins[name]+=1; pods[name][0]+=1
            elif pos==2: pods[name][1]+=1
            elif pos==3 and not is_troll4_dnf: pods[name][2]+=1
        # Ghost players: shadow ELO calc (pairwise vs lobby, no effect on others)
        if not no_ghosts:
            for pos, name, real in cup.get('ghosts', []):
                ra = ratings[name]
                k = K_BASE/(n-1)
                if gp[name] < PROV_CUPS: k *= PROV_MULT
                ghost_delta = 0.0
                for pj, nj in players:
                    e = E(ra, ratings[nj])
                    s = 1.0 if pos<pj else (0.0 if pos>pj else 0.5)
                    ghost_delta += k*(s-e)
                ratings[name] += ghost_delta
                gp[name] += 1
                history[name].append({'cup':cup['name'],'position':pos,'rating':round(ratings[name],1),'lobby_size':n})
                if pos < best[name]: best[name] = pos
                total_pos[name] += pos; avg_cups[name] += 1
                if pos==1: wins[name]+=1; pods[name][0]+=1
                elif pos==2: pods[name][1]+=1
                elif pos==3: pods[name][2]+=1
    return {'ratings': ratings, 'gp': gp, 'history': history, 'wins': wins,
            'pods': pods, 'best': best, 'total_pos': total_pos, 'avg_cups': avg_cups}

def compute_weighted_elo(cups, pct_fn=None, no_ghosts=False):
    if pct_fn is None: pct_fn = pct_mult
    w_ratings = defaultdict(lambda: STARTING)
    w_gp = defaultdict(int); w_history = defaultdict(list)
    for cup in cups:
        players = cup['players']; n = len(players)
        if n < 2: continue
        avg_field = sum(w_ratings[nm] for _, nm in players) / n
        w_cup_deltas = defaultdict(float)
        for i in range(n):
            pi, ni = players[i]; ra = w_ratings[ni]
            for j in range(n):
                if i == j: continue
                pj, nj = players[j]
                e = E(ra, w_ratings[nj])
                s = 1.0 if pi < pj else (0.0 if pi > pj else 0.5)
                win_pos = pi if pi <= pj else pj
                win_name = ni if pi <= pj else nj
                pair_quality = (w_ratings[ni] + w_ratings[nj]) / (2 * avg_field)
                k = K_BASE / (n - 1) * pct_fn(win_pos, n) * pair_quality
                if w_gp[win_name] < PROV_CUPS: k *= PROV_MULT
                w_cup_deltas[ni] += k * (s - e)
        for pos, name in players:
            w_ratings[name] += w_cup_deltas[name]
            w_gp[name] += 1
            w_history[name].append({'cup': cup['name'], 'position': pos, 'rating': round(w_ratings[name], 1), 'lobby_size': n})
        # Ghost players: shadow weighted ELO calc (pairwise vs lobby, no effect on others)
        if not no_ghosts:
            for pos, name, real in cup.get('ghosts', []):
                ra = w_ratings[name]
                ghost_delta = 0.0
                for pj, nj in players:
                    e = E(ra, w_ratings[nj])
                    s = 1.0 if pos<pj else (0.0 if pos>pj else 0.5)
                    win_pos = pos if pos <= pj else pj
                    pair_quality = (ra + w_ratings[nj]) / (2 * avg_field)
                    k = K_BASE / (n-1) * pct_fn(win_pos, n) * pair_quality
                    if w_gp[name] < PROV_CUPS: k *= PROV_MULT
                    ghost_delta += k*(s-e)
                w_ratings[name] += ghost_delta
                w_gp[name] += 1
                w_history[name].append({'cup': cup['name'], 'position': pos, 'rating': round(w_ratings[name], 1), 'lobby_size': n})
    return {'ratings': w_ratings, 'gp': w_gp, 'history': w_history}

def build_site_list(elo_data, stat_data, cups_list, min_cups=5, no_decay=False):
    rat = elo_data['ratings']; hist = elo_data['history']
    gp_d = stat_data['gp']; wins_d = stat_data['wins']; pods_d = stat_data['pods']
    best_d = stat_data['best']; total_pos_d = stat_data['total_pos']; avg_cups_d = stat_data['avg_cups']
    # Compute decay for this cup list
    total_n = len(cups_list)
    last_idx = {}
    for idx, cup in enumerate(cups_list):
        for _, name in cup['players']:
            last_idx[name] = idx
        for entry in cup.get('ghosts', []):
            last_idx[entry[1]] = idx
    def dec(rating, name):
        if no_decay: return round(rating, 1)
        missed = total_n - 1 - last_idx.get(name, 0)
        if missed <= GRACE: return round(rating, 1)
        return round(1500 + (rating - 1500) * (DECAY ** (missed - GRACE)), 1)
    out = []
    for name in rat:
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < min_cups and not has_pod: continue
        raw = round(rat[name], 1)
        act = dec(rat[name], name)
        peak = max(h['rating'] for h in hist[name]) if hist[name] else raw
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        h_list = [{'c': cup_num(h['cup']), 'r': h['rating'], 'p': h['position']} for h in hist[name]]
        out.append({
            'name': name, 'rating': raw, 'active': act,
            'cups': gp_d[name], 'wins': wins_d[name],
            'podiums': {'gold': pods_d[name][0], 'silver': pods_d[name][1], 'bronze': pods_d[name][2]},
            'avg_position': avg, 'best_finish': best_d[name] if best_d[name] < 999 else 0,
            'peak_rating': round(peak, 1), 'history': h_list
        })
    out.sort(key=lambda p: p['active'], reverse=True)
    return out[:150]

# --- Season 2026 ---
SEASON_2026_START = 126  # COTD 126 = Jan 3, 2026
season_cups = [c for c in pure_cups if cup_num(c['name']) >= SEASON_2026_START]
print(f"\n2026 season cups: {len(season_cups)} (COTD {SEASON_2026_START}+)")

# --- Compute all variants ---
print("\nComputing standard ELO (all cups)...")
std_full = compute_standard_elo(all_cups)
print("Computing weighted ELO (all cups)...")
w_full = compute_weighted_elo(all_cups)
print("Computing standard ELO (pure cups)...")
std_pure = compute_standard_elo(pure_cups)
print("Computing weighted ELO (pure cups)...")
w_pure = compute_weighted_elo(pure_cups)
print("Computing 2026 season ELO...")
season_std = compute_standard_elo(season_cups, no_ghosts=True)
season_w = compute_weighted_elo(season_cups, no_ghosts=True)

# --- Console output (from full standard) ---
ratings = w_full['ratings']; gp = std_full['gp']; history = w_full['history']
wins = std_full['wins']; pods = std_full['pods']; best = std_full['best']
total_pos = std_full['total_pos']; avg_cups = std_full['avg_cups']

lb = sorted([(n,round(ratings[n],1),gp[n],wins[n],pods[n],best[n],total_pos[n],avg_cups[n]) for n in ratings],
    key=lambda x:x[1],reverse=True)

print("\n"+"="*105)
print(f"{'#':<5}{'Player':<26}{'Elo':<9}{'Cups':<6}{'W':<4}{'Pod':<11}{'Avg':<7}{'Peak':<9}{'Best'}")
print("="*105)
for rank,(name,rating,cp,w,pd,bf,tp,ac) in enumerate(lb,1):
    peak = max(h['rating'] for h in history[name])
    avg = tp/ac if ac > 0 else 0
    print(f"{rank:<5}{name:<26}{rating:<9}{cp:<6}{w:<4}{pd[0]}/{pd[1]}/{pd[2]:<7}{avg:<7.1f}{peak:<9}{bf}")
    if rank >= 40: break

print(f"\nTotal: {len(lb)} | 5+: {sum(1 for _,_,g,_,_,_,_,_ in lb if g>=5)} | 10+: {sum(1 for _,_,g,_,_,_,_,_ in lb if g>=10)} | 20+: {sum(1 for _,_,g,_,_,_,_,_ in lb if g>=20)}")

# Save elo_results JSON
output = {
    'parameters': {'starting_rating':STARTING,'k_base':K_BASE,'provisional_cups':PROV_CUPS,'provisional_multiplier':PROV_MULT,'cups_processed':len(all_cups)},
    'leaderboard': [
        {'rank':i+1,'name':name,'rating':rating,'cups':cp,'wins':w,
         'podiums':{'gold':pd[0],'silver':pd[1],'bronze':pd[2]},
         'avg_position':round(tp/ac,1) if ac > 0 else 0,'best_finish':bf,
         'peak_rating':max(h['rating'] for h in history[name]),
         'history':history[name]}
        for i,(name,rating,cp,w,pd,bf,tp,ac) in enumerate(lb)
    ]
}
with open(_p('elo_results_75.json'),'w') as f:
    json.dump(output,f,indent=2)
print("JSON saved")

# --- Build site lists (4 variants + season) ---
std_list      = build_site_list(std_full, std_full, all_cups)
w_list        = build_site_list(w_full,   std_full, all_cups)
std_pure_list = build_site_list(std_pure, std_pure, pure_cups)
w_pure_list   = build_site_list(w_pure,   std_pure, pure_cups)
season_list   = build_site_list(season_w, season_std, season_cups, min_cups=1, no_decay=True)

# --- alldata.json (all players, compact keys, with history) ---
def build_all_list(elo_data, stat_data, cups_list, min_cups=3, no_decay=False):
    rat = elo_data['ratings']; hist = elo_data['history']
    gp_d = stat_data['gp']
    best_d = stat_data['best']
    total_pos_d = stat_data['total_pos']
    avg_cups_d = stat_data['avg_cups']
    wins_d = stat_data['wins']; pods_d = stat_data['pods']
    total_n = len(cups_list)
    last_idx = {}
    for idx, cup in enumerate(cups_list):
        for _, name in cup['players']:
            last_idx[name] = idx
        for entry in cup.get('ghosts', []):
            last_idx[entry[1]] = idx
    def dec(rating, name):
        if no_decay: return round(rating, 1)
        missed = total_n - 1 - last_idx.get(name, 0)
        if missed <= GRACE: return round(rating, 1)
        return round(1500 + (rating - 1500) * (DECAY ** (missed - GRACE)), 1)
    out = []
    for name in rat:
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < min_cups and not has_pod: continue
        raw = round(rat[name], 1)
        act = dec(rat[name], name)
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        peak = max(h['rating'] for h in hist[name]) if hist[name] else raw
        h_list = [{'c': cup_num(h['cup']), 'r': h['rating'], 'p': h['position']} for h in hist[name]]
        out.append({
            'n': name, 'a': act, 'r': raw,
            'c': gp_d[name], 'b': best_d[name] if best_d[name] < 999 else 0,
            'v': avg, 'w': wins_d[name],
            'g': pods_d[name][0], 's': pods_d[name][1], 'z': pods_d[name][2],
            'p': round(peak, 1), 'h': h_list
        })
    out.sort(key=lambda p: p['a'], reverse=True)
    return out

alldata = {
    'standard':      build_all_list(std_full, std_full, all_cups, min_cups=1),
    'weighted':      build_all_list(w_full,   std_full, all_cups, min_cups=1),
    'standard_pure': build_all_list(std_pure, std_pure, pure_cups, min_cups=1),
    'weighted_pure': build_all_list(w_pure,   std_pure, pure_cups, min_cups=1),
    'season_2026':   build_all_list(season_w, season_std, season_cups, min_cups=1, no_decay=True),
}
with open(_p('alldata.json'), 'w') as f:
    json.dump(alldata, f, separators=(',', ':'))
print(f"alldata.json written ({len(alldata['standard'])} players)")

# --- Rising.json ---
RISING_LOOKBACK_6M = 26
RISING_LOOKBACK_3M = 13
RISING_MIN_RATING = 1600
RISING_TOP_N = 50
current_cup_n = cup_num(all_cups[-1]['name'])
lookback_6m = current_cup_n - RISING_LOOKBACK_6M
lookback_3m = current_cup_n - RISING_LOOKBACK_3M

def build_rising(player_list, lookback_cup):
    results = []
    for p in player_list:
        if p['rating'] < RISING_MIN_RATING:
            continue
        past = [h for h in p['history'] if h['c'] <= lookback_cup]
        if not past:
            continue
        past_r = past[-1]['r']
        pct = round((p['rating'] - past_r) / past_r * 100, 1)
        if pct < 1.0:
            continue
        results.append({
            'name':        p['name'],
            'rating_now':  p['rating'],
            'rating_then': round(past_r, 1),
            'pct':         pct,
        })
    results.sort(key=lambda x: x['pct'], reverse=True)
    return results[:50]

def build_rising_combined(player_list):
    """Anyone with >=1% growth in either 6M or 3M gets included.
       Eligible: rating >= 1600 OR in top 50 by rating."""
    top50 = set(p['name'] for p in sorted(player_list, key=lambda x: x['active'], reverse=True)[:RISING_TOP_N])
    entries = {}
    for p in player_list:
        if p['rating'] < RISING_MIN_RATING and p['name'] not in top50:
            continue
        past6 = [h for h in p['history'] if h['c'] <= lookback_6m]
        past3 = [h for h in p['history'] if h['c'] <= lookback_3m]
        if past6:
            r_then6 = past6[-1]['r']
        else:
            # No 6M history — use 1500 starting ELO
            r_then6 = 1500.0
        pct6 = round((p['rating'] - r_then6) / r_then6 * 100, 1)
        r_then = round(r_then6, 1)
        r_then3 = past3[-1]['r'] if past3 else 1500.0
        pct3 = round((p['rating'] - r_then3) / r_then3 * 100, 1)
        if pct6 >= 1.0 or pct3 >= 1.0:
            entries[p['name']] = {
                'name': p['name'], 'rating_now': p['rating'],
                'rating_then': r_then, 'pct': pct6, 'pct_3m': pct3,
            }
    return sorted(entries.values(), key=lambda x: x['pct'], reverse=True)[:50]

rising_out = {
    'current_cup':   current_cup_n,
    'lookback_cup':  lookback_6m,
    'lookback_3m':   lookback_3m,
    'lookback_cups': RISING_LOOKBACK_6M,
    'min_rating':    RISING_MIN_RATING,
    'standard':      build_rising_combined(std_list),
    'weighted':      build_rising_combined(w_list),
    'standard_pure': build_rising_combined(std_pure_list),
    'weighted_pure': build_rising_combined(w_pure_list),
}
with open(_p('rising.json'), 'w') as f:
    json.dump(rising_out, f, indent=2)
print("rising.json written")

# --- Lexer Curse ELO ---
# Remap positions: 4th=best, spiraling outward (4,3,5,2,6,1,7,8,9...)
def curse_remap(pos):
    return {4:1, 3:2, 5:3, 2:4, 6:5, 1:6}.get(pos, pos)

def curse_pct_mult(pos, n):
    """Fixed multipliers by curse position (real pos in parens)."""
    if pos == 1: return 5.0    # real 4th
    if pos <= 3: return 2.5    # real 3rd, 5th
    if pos <= 5: return 1.0    # real 2nd, 6th
    if pos <= 7: return 0.5    # real 1st, 7th
    return 0.3

curse_cups = []
for cup in all_cups:
    cp = [(curse_remap(pos), name) for pos, name in cup['players']]
    cp.sort(key=lambda x: x[0])
    curse_cups.append({'name': cup['name'], 'players': cp})

# Count real 4th-place finishes from original data
fourths_count = defaultdict(int)
for cup in all_cups:
    for pos, name in cup['players']:
        if pos == 4: fourths_count[name] += 1

print("\nComputing Lexer Curse ELO (weighted, 4th-biased)...")
curse_result = compute_weighted_elo(curse_cups, pct_fn=curse_pct_mult)

# Precompute last cup index for decay
curse_last_idx = {}
for idx, cup in enumerate(all_cups):
    for _, name in cup['players']:
        curse_last_idx[name] = idx

curse_list = []
for name in curse_result['ratings']:
    if curse_result['gp'][name] < 3: continue
    raw = round(curse_result['ratings'][name], 1)
    missed = len(all_cups) - 1 - curse_last_idx.get(name, 0)
    active = round(1500 + (raw - 1500) * (DECAY ** (missed - GRACE)), 1) if missed > GRACE else raw
    h = curse_result['history'][name]
    peak = max(e['rating'] for e in h) if h else raw
    curse_list.append({
        'name': name, 'rating': raw, 'active': active,
        'cups': curse_result['gp'][name],
        'fourths': fourths_count[name],
        'peak_rating': round(peak, 1),
        'history': [{'c': cup_num(e['cup']), 'r': e['rating'], 'p': e['position']} for e in h]
    })
curse_list.sort(key=lambda p: p['active'], reverse=True)
curse_list = curse_list[:150]

with open(_p('lexercurse.json'), 'w') as f:
    json.dump({'l': curse_list}, f, separators=(',', ':'))
print(f"lexercurse.json written ({len(curse_list)} players)")
