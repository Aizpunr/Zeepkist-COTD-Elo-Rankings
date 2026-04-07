"""
build_whatif.py — What-If ELO: recalculate ratings with early rounds removed.

For each skip level N, removes all players eliminated in rounds 1..N from every cup,
renumbers positions, and recalculates weighted ELO from scratch.

Outputs whatif.json for the whatif.html frontend.
"""
import openpyxl, json, re, sys, os
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

base = os.path.dirname(os.path.abspath(__file__))
def _p(f): return os.path.join(base, f)

# ── Aliases (same pattern as build_fastest.py) ──────────────────────────────

def load_aliases():
    name_map = {}
    canonical = {}
    lines = open(_p('elo_75.py'), encoding='utf-8').readlines()
    collecting = False
    buf = []
    for line in lines:
        if not collecting and re.match(r'^CANONICAL\s*=\s*\{', line):
            collecting = True
        if collecting:
            buf.append(line)
            if line.strip() == '}':
                break
    if buf:
        block = ''.join(buf).split('=', 1)[1].strip()
        canonical = eval(block)
        for canon, aliases in canonical.items():
            for alias in aliases:
                name_map[alias] = canon
    return name_map, canonical

NAME_MAP, CANONICAL = load_aliases()

def strip_tag(name):
    return re.sub(r'\[.*?\]\s*', '', name).strip()

def normalize(name):
    if name in NAME_MAP:
        return NAME_MAP[name]
    stripped = strip_tag(name)
    if stripped in NAME_MAP:
        return NAME_MAP[stripped]
    if stripped != name and stripped in CANONICAL:
        return stripped
    return name

# ── Parser (reads Position + Name + Elim Round) ────────────────────────────

def parse_file_with_rounds(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except FileNotFoundError:
        print(f"  SKIP: {filepath} not found")
        return []
    cups = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if val != 'Position':
                    continue
                # Check for round column
                has_round = (ci + 3 < len(rows[ri]) and
                             rows[ri][ci + 3] in ('Elim Round', 'Round'))
                if not has_round:
                    continue
                # Find cup name above
                cup_name = None
                for sr in range(ri - 1, max(ri - 5, -1), -1):
                    v = rows[sr][ci] if ci < len(rows[sr]) else None
                    if v and (str(v).startswith('COTD') or str(v).startswith('COTW')):
                        cup_name = str(v).strip()
                        break
                if not cup_name:
                    continue
                # Read players
                players = []
                last_pos = None
                for row2 in rows[ri + 1:]:
                    if ci >= len(row2) or ci + 1 >= len(row2):
                        continue
                    pos, name = row2[ci], row2[ci + 1]
                    if name is None:
                        break
                    name_str = str(name).strip()
                    if name_str.startswith('*'):
                        continue
                    rnd_val = row2[ci + 3] if ci + 3 < len(row2) else None
                    round_num = None
                    if rnd_val is not None:
                        try:
                            round_num = int(float(str(rnd_val)))
                        except (ValueError, TypeError):
                            pass
                    if pos is not None:
                        try:
                            last_pos = int(float(str(pos).rstrip('*').strip()))
                            players.append((last_pos, name_str, round_num))
                        except (ValueError, TypeError):
                            continue
                    elif last_pos is not None:
                        players.append((last_pos, name_str, round_num))
                if players:
                    cups.append({'name': cup_name,
                                 'players': sorted(players, key=lambda x: x[0])})
    return cups

# ── Load cups ───────────────────────────────────────────────────────────────

# Read xlsx file list from elo_75.py source (auto-syncs when files change)
elo_src = open(_p('elo_75.py'), encoding='utf-8').read()
xlsx_files = re.findall(r"parse_file\(_p\('(.+?\.xlsx)'\)\)", elo_src)
xlsx_files = [f for f in xlsx_files if 'Troll' not in f and 'roulette' not in f.lower()]

all_cups = []
for f in xlsx_files:
    parsed = parse_file_with_rounds(_p(f))
    all_cups += parsed
    print(f"  {f}: {len(parsed)} cups")

def cup_num(name):
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else 0

all_cups.sort(key=lambda c: cup_num(c['name']))

seen = set()
deduped = []
for c in all_cups:
    n = cup_num(c['name'])
    if n not in seen:
        seen.add(n)
        deduped.append(c)
all_cups = deduped

# COTD 16 fix: both finalists DNF'd — tied at 2nd
for c in all_cups:
    if c['name'] == 'COTD 16':
        c['players'] = [(2 if pos == 1 else pos, name, rnd)
                        for pos, name, rnd in c['players']]
        break

print(f"\nLoaded {len(all_cups)} cups with round data")

# ── Normalize names + ghosts ────────────────────────────────────────────────

all_names = set()
for cup in all_cups:
    for _, name, _ in cup['players']:
        all_names.add(name)

for n in all_names:
    stripped = strip_tag(n)
    if stripped != n and stripped in CANONICAL and n not in NAME_MAP:
        NAME_MAP[n] = stripped

for cup in all_cups:
    cup['players'] = [(pos, normalize(name), rnd) for pos, name, rnd in cup['players']]

# Ghost accounts: "account (elo=RealPlayer)" → real player gets credit
for cup in all_cups:
    new_players = []
    for pos, name, rnd in cup['players']:
        m = re.match(r'^(.+?)\s*\(elo=(.+?)\)$', name)
        if m:
            new_players.append((pos, normalize(m.group(2).strip()), rnd))
        else:
            new_players.append((pos, name, rnd))
    cup['players'] = new_players

unique = set(name for cup in all_cups for _, name, _ in cup['players'])
print(f"Unique players: {len(unique)}")

# ── ELO functions (from elo_75.py, no ghost handling) ───────────────────────

STARTING = 1500
K_BASE = 32
PROV_CUPS = 12
PROV_MULT = 1.5
DECAY = 0.995
GRACE = 3

def E(ra, rb):
    return 1.0 / (1.0 + 10.0**((rb - ra) / 400.0))

def pct_mult(pos, n):
    pct = pos / n
    if pct <= 0.08: return 3.0
    if pct <= 0.15: return 2.0
    if pct <= 0.25: return 1.3
    if pct <= 0.50: return 0.8
    return 0.5

def compute_standard_elo(cups):
    ratings = defaultdict(lambda: STARTING)
    gp = defaultdict(int)
    wins = defaultdict(int)
    pods = defaultdict(lambda: [0, 0, 0])
    best = defaultdict(lambda: 999)
    total_pos = defaultdict(int)
    avg_cups = defaultdict(int)
    for cup in cups:
        players = cup['players']; n = len(players)
        if n < 2: continue
        deltas = defaultdict(float)
        for i in range(n):
            pi, ni = players[i]
            k = K_BASE / (n - 1)
            if gp[ni] < PROV_CUPS: k *= PROV_MULT
            for j in range(n):
                if i == j: continue
                pj, nj = players[j]
                s = 1.0 if pi < pj else (0.0 if pi > pj else 0.5)
                deltas[ni] += k * (s - E(ratings[ni], ratings[nj]))
        for pos, name in players:
            ratings[name] += deltas[name]
            gp[name] += 1
            total_pos[name] += pos
            avg_cups[name] += 1
            if pos < best[name]: best[name] = pos
            if pos == 1: wins[name] += 1; pods[name][0] += 1
            elif pos == 2: pods[name][1] += 1
            elif pos == 3: pods[name][2] += 1
    return {'ratings': ratings, 'gp': gp, 'wins': wins, 'pods': pods,
            'best': best, 'total_pos': total_pos, 'avg_cups': avg_cups}

def compute_weighted_elo(cups):
    ratings = defaultdict(lambda: STARTING)
    gp = defaultdict(int)
    peak = defaultdict(lambda: STARTING)
    for cup in cups:
        players = cup['players']; n = len(players)
        if n < 2: continue
        avg_field = sum(ratings[nm] for _, nm in players) / n
        deltas = defaultdict(float)
        for i in range(n):
            pi, ni = players[i]; ra = ratings[ni]
            for j in range(n):
                if i == j: continue
                pj, nj = players[j]
                s = 1.0 if pi < pj else (0.0 if pi > pj else 0.5)
                win_pos = pi if pi <= pj else pj
                win_name = ni if pi <= pj else nj
                pair_quality = (ratings[ni] + ratings[nj]) / (2 * avg_field)
                k = K_BASE / (n - 1) * pct_mult(win_pos, n) * pair_quality
                if gp[win_name] < PROV_CUPS: k *= PROV_MULT
                deltas[ni] += k * (s - E(ra, ratings[nj]))
        for pos, name in players:
            ratings[name] += deltas[name]
            gp[name] += 1
            if ratings[name] > peak[name]:
                peak[name] = ratings[name]
    return {'ratings': ratings, 'gp': gp, 'peak': peak}

# ── Round filtering ─────────────────────────────────────────────────────────

def filter_cup(cup, skip_n):
    """Remove players eliminated in rounds 1..skip_n, renumber positions."""
    if skip_n == 0:
        return {'name': cup['name'],
                'players': [(pos, name) for pos, name, _ in cup['players']]}

    remaining = [(pos, name, rnd) for pos, name, rnd in cup['players']
                 if rnd is None or rnd > skip_n]
    if len(remaining) < 2:
        return None

    # Group by original position (preserves ties), renumber sequentially
    new_players = []
    new_pos = 1
    i = 0
    while i < len(remaining):
        orig_pos = remaining[i][0]
        group = []
        while i < len(remaining) and remaining[i][0] == orig_pos:
            group.append(remaining[i][1])
            i += 1
        for name in group:
            new_players.append((new_pos, name))
        new_pos += len(group)
    return {'name': cup['name'], 'players': new_players}

# ── Build level list ────────────────────────────────────────────────────────

def build_level_list(w_data, std_data, cups_list):
    rat = w_data['ratings']
    gp_d = std_data['gp']
    wins_d = std_data['wins']
    pods_d = std_data['pods']
    best_d = std_data['best']
    total_pos_d = std_data['total_pos']
    avg_cups_d = std_data['avg_cups']
    peak_d = w_data['peak']

    total_n = len(cups_list)
    last_idx = {}
    for idx, cup in enumerate(cups_list):
        for _, name in cup['players']:
            last_idx[name] = idx

    def dec(rating, name):
        missed = total_n - 1 - last_idx.get(name, 0)
        if missed <= GRACE: return round(rating, 1)
        return round(1500 + (rating - 1500) * (DECAY ** (missed - GRACE)), 1)

    out = []
    for name in rat:
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < 5 and not has_pod: continue
        raw = round(rat[name], 1)
        act = dec(rat[name], name)
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        out.append({
            'n': name, 'a': act, 'r': raw,
            'c': gp_d[name], 'b': best_d[name] if best_d[name] < 999 else 0,
            'v': avg, 'w': wins_d[name],
            'g': pods_d[name][0], 's': pods_d[name][1], 'z': pods_d[name][2],
            'p': round(peak_d.get(name, raw), 1),
        })
    out.sort(key=lambda p: p['a'], reverse=True)
    return out

# ── Auto-detect max skip ───────────────────────────────────────────────────

def count_survivors(cup, skip_n):
    return sum(1 for _, _, rnd in cup['players'] if rnd is None or rnd > skip_n)

max_skip = 0
for test in range(1, 25):
    viable = sum(1 for cup in all_cups if count_survivors(cup, test) >= 5)
    if viable < 10:
        max_skip = test - 1
        break
    max_skip = test

print(f"Max skip: {max_skip}\n")
for s in range(max_skip + 1):
    viable = sum(1 for c in all_cups if count_survivors(c, s) >= 2)
    avg_sz = sum(count_survivors(c, s) for c in all_cups) / len(all_cups)
    print(f"  skip {s:2d}: {viable:3d} cups, avg {avg_sz:4.0f} players")

# ── Compute ─────────────────────────────────────────────────────────────────

levels = {}
meta_cups = {}

for skip in range(max_skip + 1):
    filtered = []
    for cup in all_cups:
        fc = filter_cup(cup, skip)
        if fc and len(fc['players']) >= 2:
            filtered.append(fc)
    meta_cups[str(skip)] = len(filtered)

    std = compute_standard_elo(filtered)
    w = compute_weighted_elo(filtered)
    level_list = build_level_list(w, std, filtered)
    levels[str(skip)] = level_list

    top3 = ', '.join(f"{p['n']} ({p['a']})" for p in level_list[:3])
    print(f"\n  skip {skip}: {len(filtered)} cups, {len(level_list)} players — {top3}")

# ── Output ──────────────────────────────────────────────────────────────────

output = {
    'max_skip': max_skip,
    'levels': levels,
    'meta': {'cups_per_level': meta_cups},
}
with open(_p('whatif.json'), 'w') as f:
    json.dump(output, f, separators=(',', ':'))

sz = os.path.getsize(_p('whatif.json'))
print(f"\nwhatif.json written ({sz // 1024} KB, skip 0-{max_skip})")
