"""
build_altrank.py — Alternative Rankings: Standard ELO + TrueSkill.

Standard ELO is the legacy main-page metric, now demoted to alt-rank since
index.html runs on Weighted ELO + Glicko-2 ("Skill Ranking"). TrueSkill is
the second alt-rank view. Glicko-2 moved to elo_engine.py since it powers
the main page.

Both systems are computed from scratch — no pip dependencies, pure Python.

Reads alldata.json (written by elo_engine.py) and appends:
  - 'standard', 'standard_pure', 'trueskill', 'trueskill_pure', 'cupDates'

Reads rising.json (written by elo_engine.py) and appends:
  - 'standard', 'standard_pure'
"""
import openpyxl, json, re, sys, os, math
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

base = os.path.dirname(os.path.abspath(__file__))
def _p(f): return os.path.join(base, f)

# ── Aliases (eval from elo_engine.py source) ──────────────────────────────

def load_aliases():
    name_map = {}
    canonical = {}
    lines = open(_p('elo_engine.py'), encoding='utf-8').readlines()
    collecting = False
    buf = []
    for line in lines:
        if not collecting and re.match(r'^CANONICAL\s*=\s*\{', line):
            collecting = True
        if collecting:
            buf.append(line)
            if line.strip() == '}':
                break
    if buf:
        block = ''.join(buf).split('=', 1)[1].strip()
        canonical = eval(block)
        for canon, aliases in canonical.items():
            for alias in aliases:
                name_map[alias] = canon
    return name_map, canonical

NAME_MAP, CANONICAL = load_aliases()

def strip_tag(name):
    return re.sub(r'\[.*?\]\s*', '', name).strip()

def normalize(name):
    if name in NAME_MAP:
        return NAME_MAP[name]
    stripped = strip_tag(name)
    if stripped in NAME_MAP:
        return NAME_MAP[stripped]
    if stripped != name and stripped in CANONICAL:
        return stripped
    return name

# ── Parsers (from elo_engine.py) ──────────────────────────────────────────

def parse_file(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except FileNotFoundError:
        print(f"  SKIP: {filepath} not found")
        return []
    cups = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        position_cells = []
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if val == 'Position':
                    position_cells.append((ri, ci))
        for pos_row, pos_col in position_cells:
            cup_name = None
            for sr in range(pos_row - 1, max(pos_row - 5, -1), -1):
                val = rows[sr][pos_col] if pos_col < len(rows[sr]) else None
                if val and (str(val).startswith('COTD') or str(val).startswith('COTW')):
                    cup_name = str(val).strip()
                    break
            if not cup_name: continue
            players = []
            last_pos = None
            for row in rows[pos_row + 1:]:
                if pos_col >= len(row) or pos_col + 1 >= len(row): continue
                pos, name = row[pos_col], row[pos_col + 1]
                if name is None: continue
                name_str = str(name).strip()
                if name_str.startswith('*'): continue
                if pos is not None:
                    try:
                        pos_clean = str(pos).rstrip('*').strip()
                        last_pos = int(float(pos_clean))
                        players.append((last_pos, name_str))
                    except: continue
                elif last_pos is not None:
                    players.append((last_pos, name_str))
            if players:
                cups.append({'name': cup_name, 'players': sorted(players, key=lambda x: x[0])})
    return cups

def parse_troll_cups(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Troll COTDs']
    rows = list(ws.iter_rows(values_only=True))
    troll_defs = [
        ('Troll COTD 1', 15.5, 0), ('Troll COTD 2', 26.5, 6),
        ('Troll COTD 3', 36.5, 12), ('Troll COTD 4', 41.5, 17),
        ('Troll COTD 5', 44.5, 23), ('Troll COTD 6', 48.5, 29),
        ('Troll COTD 7', 50.5, 35), ('Troll COTD 8', 56.5, 41),
        ('Troll COTD 9', 63.5, 47), ('Troll COTD 10', 71.5, 53),
        ('Troll COTD 11', 88.5, 59),
    ]
    troll4_override = [
        (1, 'justMaki'), (2, 'Mark'), (3, 'Lexer'),
        (3, '[CTR]R0nanC'), (3, 'ButItRuns'), (3, 'St Nicholas'), (3, 'Hi Im Yolo'),
        (3, 'RoundNzt'), (3, '[CTR]L3it3R'), (3, 'FINSTER83'), (3, 'Neb'),
        (3, '[CTR]Hydro'), (3, '[RTR]Fwogiie'), (3, '[heyo]Mr. Hubub'), (3, 'agix'),
        (3, 'koz'), (3, '[6dog] schmxrg'), (3, 'Lynhardt'),
    ]
    cups = []
    for name, order, pc in troll_defs:
        players = []
        for row in rows[5:]:
            if pc >= len(row) or pc+1 >= len(row): continue
            pos, nm = row[pc], row[pc+1]
            if pos is None or nm is None: continue
            nm = str(nm).strip()
            if 'Other Player' in nm: continue
            nm = nm.replace('[CTR[', '[CTR]').replace('[RTR[', '[RTR]')
            try: players.append((int(float(pos)), nm))
            except: continue
        if len(players) >= 2:
            cups.append({'name': name, 'players': sorted(players, key=lambda x: x[0])})
        if name == 'Troll COTD 4':
            cups[-1 if cups and cups[-1]['name'] == 'Troll COTD 4' else len(cups):] = []
            cups.append({'name': name, 'players': troll4_override})
    return cups

# ── Cup loading + normalization ───────────────────────────────────────────

# Read xlsx file list from elo_engine.py source
elo_src = open(_p('elo_engine.py'), encoding='utf-8').read()
xlsx_files = re.findall(r"parse_file\(_p\('(.+?\.xlsx)'\)\)", elo_src)
xlsx_files = [f for f in xlsx_files if 'Troll' not in f and 'roulette' not in f.lower()]

print("Loading cups...")
all_cups = []
for xlsx in xlsx_files:
    all_cups += parse_file(_p(xlsx))

# Roulette + troll cups
roulette_files = [f for f in re.findall(r"parse_file\(_p\('(.+?\.xlsx)'\)\)", elo_src)
                  if 'roulette' in f.lower()]
for f in roulette_files:
    all_cups += parse_file(_p(f))

troll_files = re.findall(r"parse_troll_cups\(_p\('(.+?\.xlsx)'\)\)", elo_src)
for f in troll_files:
    all_cups += parse_troll_cups(_p(f))

SPECIAL_CUP_ORDER = {
    'COTD Roulette 1': 25.5, 'COTD Roulette 2': 65.5,
    'Troll COTD 1': 15.5, 'Troll COTD 2': 26.5,
    'Troll COTD 3': 36.5, 'Troll COTD 4': 41.5,
    'Troll COTD 5': 44.5, 'Troll COTD 6': 48.5,
    'Troll COTD 7': 50.5, 'Troll COTD 8': 56.5,
    'Troll COTD 9': 63.5, 'Troll COTD 10': 71.5,
    'Troll COTD 11': 88.5,
}

def cup_num(name):
    if name in SPECIAL_CUP_ORDER: return SPECIAL_CUP_ORDER[name]
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else 0

all_cups.sort(key=lambda c: cup_num(c['name']))

# Deduplicate
seen = set()
deduped = []
for c in all_cups:
    n = cup_num(c['name'])
    if n not in seen:
        seen.add(n)
        deduped.append(c)
all_cups = deduped

def is_nonstandard(cup_name):
    return cup_name.startswith('Troll ') or 'Roulette' in cup_name

pure_cups = [c for c in all_cups if not is_nonstandard(c['name'])]
print(f"Loaded {len(all_cups)} cups ({len(pure_cups)} pure)")

# COTD 16 fix: both finalists DNF'd — tied at 2nd
for c in all_cups:
    if c['name'] == 'COTD 16':
        c['players'] = [(2 if pos == 1 else pos, name) for pos, name in c['players']]
        break

# Normalize names
for cup in all_cups:
    cup['players'] = [(pos, normalize(name)) for pos, name in cup['players']]

# Ghost handling: "account_name (elo=RealPlayer)"
for cup in all_cups:
    new_players = []
    for pos, name in cup['players']:
        m = re.match(r'^(.+?)\s*\(elo=(.+?)\)$', name)
        if m:
            real_name = normalize(m.group(2).strip())
            new_players.append((pos, real_name))
        else:
            new_players.append((pos, name))
    cup['players'] = new_players

unique = set(name for cup in all_cups for _, name in cup['players'])
print(f"Unique players: {len(unique)}")

# ═══════════════════════════════════════════════════════════════════════════
# TRUESKILL — from scratch
# ═══════════════════════════════════════════════════════════════════════════

# Gaussian math
def norm_pdf(x):
    return math.exp(-x * x / 2) / math.sqrt(2 * math.pi)

def norm_cdf(x):
    return 0.5 * math.erfc(-x / math.sqrt(2))

def v_func(t, eps=0):
    """Truncated Gaussian mean correction (win case)."""
    denom = norm_cdf(t - eps)
    if denom < 1e-10:
        return -t + eps
    return norm_pdf(t - eps) / denom

def w_func(t, eps=0):
    """Truncated Gaussian variance correction (win case)."""
    v = v_func(t, eps)
    return v * (v + t - eps)

# Constants
TS_MU0 = 25.0
TS_SIGMA0 = TS_MU0 / 3          # 8.333
TS_BETA = TS_SIGMA0 / 2          # 4.167 — performance variance
TS_TAU = TS_SIGMA0 / 100         # 0.0833 — dynamics factor (sigma drift per cup)
TS_SIGMA_FLOOR = 0.5             # minimum sigma to prevent collapse
TS_DISPLAY_SCALE = 40.0          # maps native → 1500 range
TS_DISPLAY_BASE = 1500.0

def ts_display(mu, sigma):
    """Scale TrueSkill conservative rating (mu - 3σ) to 1500 range."""
    return round(TS_DISPLAY_BASE + (mu - 3 * sigma) * TS_DISPLAY_SCALE, 1)

def compute_trueskill(cups):
    ts = {}  # name → (mu, sigma)
    gp = defaultdict(int)
    history = defaultdict(list)
    wins = defaultdict(int)
    pods = defaultdict(lambda: [0, 0, 0])
    best = defaultdict(lambda: 999)
    total_pos = defaultdict(int)
    avg_cups = defaultdict(int)
    peak = defaultdict(lambda: -9999)

    for cup in cups:
        players = cup['players']
        n = len(players)
        if n < 2: continue

        # Sigma drift for all known players (inactivity increases uncertainty)
        for name in ts:
            mu, sigma = ts[name]
            ts[name] = (mu, min(math.sqrt(sigma**2 + TS_TAU**2), TS_SIGMA0))

        # Initialize new players
        for _, name in players:
            if name not in ts:
                ts[name] = (TS_MU0, TS_SIGMA0)

        # All pairwise updates, scaled by 1/(N-1)
        delta_mu = defaultdict(float)
        delta_sig2 = defaultdict(float)

        for i in range(n):
            pi, ni = players[i]
            mu_i, sig_i = ts[ni]
            for j in range(n):
                if i == j: continue
                pj, nj = players[j]
                mu_j, sig_j = ts[nj]
                c = math.sqrt(2 * TS_BETA**2 + sig_i**2 + sig_j**2)
                t = (mu_i - mu_j) / c

                if pi < pj:       # i beat j
                    v = v_func(t)
                    w = w_func(t)
                elif pi > pj:     # j beat i
                    v = -v_func(-t)
                    w = w_func(-t)
                else:             # tie — skip
                    continue

                delta_mu[ni] += (sig_i**2 / c) * v / (n - 1)
                delta_sig2[ni] -= (sig_i**4 / c**2) * w / (n - 1)

        # Apply updates + track stats
        for pos, name in players:
            mu, sigma = ts[name]
            mu += delta_mu[name]
            sigma = math.sqrt(max(sigma**2 + delta_sig2[name], TS_SIGMA_FLOOR**2))
            ts[name] = (mu, sigma)

            display = ts_display(mu, sigma)
            gp[name] += 1
            history[name].append({'cup': cup['name'], 'position': pos,
                                  'rating': display, 'lobby_size': n})
            if display > peak[name]:
                peak[name] = display

            is_troll4_dnf = cup['name'] == 'Troll COTD 4' and pos == 3
            if not is_troll4_dnf:
                total_pos[name] += pos
                avg_cups[name] += 1
                if pos < best[name]: best[name] = pos
            if pos == 1: wins[name] += 1; pods[name][0] += 1
            elif pos == 2: pods[name][1] += 1
            elif pos == 3 and not is_troll4_dnf: pods[name][2] += 1

    # Build ratings + uncertainty dicts
    ratings = {}
    uncertainty = {}
    for name in ts:
        mu, sigma = ts[name]
        ratings[name] = ts_display(mu, sigma)
        uncertainty[name] = round(sigma * TS_DISPLAY_SCALE, 1)

    return {'ratings': ratings, 'uncertainty': uncertainty,
            'gp': gp, 'history': history, 'wins': wins,
            'pods': pods, 'best': best, 'total_pos': total_pos, 'avg_cups': avg_cups,
            'peak': peak}


# ═══════════════════════════════════════════════════════════════════════════
# STANDARD ELO — the original pairwise system; lives here now because the
# main page uses Weighted ELO + Glicko-2. Standard is shown only on altrank.
# ═══════════════════════════════════════════════════════════════════════════

STARTING = 1500
K_BASE = 32
PROV_CUPS = 12
PROV_MULT = 1.5
DECAY = 0.995
GRACE = 3
RISING_LOOKBACK_6M = 26
RISING_LOOKBACK_3M = 13
RISING_MIN_RATING = 1600
RISING_TOP_N = 50

def E(ra, rb): return 1.0 / (1.0 + 10.0 ** ((rb - ra) / 400.0))

def compute_standard_elo(cups, no_ghosts=False):
    ratings = defaultdict(lambda: STARTING)
    gp = defaultdict(int); history = defaultdict(list)
    wins = defaultdict(int); pods = defaultdict(lambda: [0, 0, 0])
    best = defaultdict(lambda: 999); total_pos = defaultdict(int); avg_cups = defaultdict(int)
    for cup in cups:
        players = cup['players']; n = len(players)
        if n < 2: continue
        deltas = defaultdict(float)
        for i in range(n):
            pi, ni = players[i]; ra = ratings[ni]
            k = K_BASE / (n - 1)
            if gp[ni] < PROV_CUPS: k *= PROV_MULT
            for j in range(n):
                if i == j: continue
                pj, nj = players[j]
                e = E(ra, ratings[nj])
                s = 1.0 if pi < pj else (0.0 if pi > pj else 0.5)
                deltas[ni] += k * (s - e)
        for pos, name in players:
            ratings[name] += deltas[name]; gp[name] += 1
            history[name].append({'cup': cup['name'], 'position': pos,
                                  'rating': round(ratings[name], 1), 'lobby_size': n})
            is_troll4_dnf = cup['name'] == 'Troll COTD 4' and pos == 3
            if not is_troll4_dnf:
                total_pos[name] += pos
                avg_cups[name] += 1
                if pos < best[name]: best[name] = pos
            if pos == 1: wins[name] += 1; pods[name][0] += 1
            elif pos == 2: pods[name][1] += 1
            elif pos == 3 and not is_troll4_dnf: pods[name][2] += 1
        # Ghost players: shadow ELO (pairwise vs lobby, no effect on others)
        if not no_ghosts:
            for pos, name, real in cup.get('ghosts', []):
                ra = ratings[name]
                k = K_BASE / (n - 1)
                if gp[name] < PROV_CUPS: k *= PROV_MULT
                ghost_delta = 0.0
                for pj, nj in players:
                    e = E(ra, ratings[nj])
                    s = 1.0 if pos < pj else (0.0 if pos > pj else 0.5)
                    ghost_delta += k * (s - e)
                ratings[name] += ghost_delta
                gp[name] += 1
                history[name].append({'cup': cup['name'], 'position': pos,
                                      'rating': round(ratings[name], 1), 'lobby_size': n})
                if pos < best[name]: best[name] = pos
                total_pos[name] += pos; avg_cups[name] += 1
                if pos == 1: wins[name] += 1; pods[name][0] += 1
                elif pos == 2: pods[name][1] += 1
                elif pos == 3: pods[name][2] += 1
    return {'ratings': ratings, 'gp': gp, 'history': history, 'wins': wins,
            'pods': pods, 'best': best, 'total_pos': total_pos, 'avg_cups': avg_cups}

# ── Site-list + rising helpers (ported from elo_engine.py for the standard
# rising.json keys) ─────────────────────────────────────────────────────────

def build_site_list(elo_data, cups_list, min_cups=5, no_decay=False):
    rat = elo_data['ratings']; hist = elo_data['history']
    gp_d = elo_data['gp']; wins_d = elo_data['wins']; pods_d = elo_data['pods']
    best_d = elo_data['best']; total_pos_d = elo_data['total_pos']; avg_cups_d = elo_data['avg_cups']
    total_n = len(cups_list)
    last_idx = {}
    for idx, cup in enumerate(cups_list):
        for _, name in cup['players']:
            last_idx[name] = idx
        for entry in cup.get('ghosts', []):
            last_idx[entry[1]] = idx
    def dec(rating, name):
        if no_decay: return round(rating, 1)
        missed = total_n - 1 - last_idx.get(name, 0)
        if missed <= GRACE: return round(rating, 1)
        return round(1500 + (rating - 1500) * (DECAY ** (missed - GRACE)), 1)
    out = []
    for name in rat:
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < min_cups and not has_pod: continue
        raw = round(rat[name], 1)
        act = dec(rat[name], name)
        peak = max(h['rating'] for h in hist[name]) if hist[name] else raw
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        h_list = [{'c': cup_num(h['cup']), 'r': h['rating'], 'p': h['position']} for h in hist[name]]
        out.append({
            'name': name, 'rating': raw, 'active': act,
            'cups': gp_d[name], 'wins': wins_d[name],
            'podiums': {'gold': pods_d[name][0], 'silver': pods_d[name][1], 'bronze': pods_d[name][2]},
            'avg_position': avg, 'best_finish': best_d[name] if best_d[name] < 999 else 0,
            'peak_rating': round(peak, 1), 'history': h_list,
        })
    out.sort(key=lambda p: p['active'], reverse=True)
    return out[:150]

def build_rising_combined(player_list, lookback_6m, lookback_3m):
    """Anyone with >=1% growth in either 6M or 3M gets included.
       Eligible: rating >= RISING_MIN_RATING OR in top RISING_TOP_N by active."""
    top50 = set(p['name'] for p in sorted(player_list, key=lambda x: x['active'], reverse=True)[:RISING_TOP_N])
    entries = {}
    for p in player_list:
        if p['rating'] < RISING_MIN_RATING and p['name'] not in top50:
            continue
        past6 = [h for h in p['history'] if h['c'] <= lookback_6m]
        past3 = [h for h in p['history'] if h['c'] <= lookback_3m]
        r_then6 = past6[-1]['r'] if past6 else 1500.0
        pct6 = round((p['rating'] - r_then6) / r_then6 * 100, 1)
        r_then = round(r_then6, 1)
        r_then3 = past3[-1]['r'] if past3 else 1500.0
        pct3 = round((p['rating'] - r_then3) / r_then3 * 100, 1)
        if pct6 >= 1.0 or pct3 >= 1.0:
            entries[p['name']] = {
                'name': p['name'], 'rating_now': p['rating'],
                'rating_then': r_then, 'pct': pct6, 'pct_3m': pct3,
            }
    return sorted(entries.values(), key=lambda x: x['pct'], reverse=True)[:50]


# ═══════════════════════════════════════════════════════════════════════════
# OUTPUT
# ═══════════════════════════════════════════════════════════════════════════

DECAY = 0.995
GRACE = 3

def build_all_list(elo_data, cups_list, min_cups=1):
    """Compact-key per-player entries for alldata.json. Supports both
    Standard ELO (no 'uncertainty'/'peak' in elo_data) and TrueSkill
    (both present). 'u' field included only when uncertainty is present."""
    rat = elo_data['ratings']
    unc_d = elo_data.get('uncertainty')
    peak_d = elo_data.get('peak')
    hist = elo_data['history']
    gp_d = elo_data['gp']
    best_d = elo_data['best']
    total_pos_d = elo_data['total_pos']
    avg_cups_d = elo_data['avg_cups']
    wins_d = elo_data['wins']
    pods_d = elo_data['pods']

    total_n = len(cups_list)
    last_idx = {}
    for idx, cup in enumerate(cups_list):
        for _, name in cup['players']:
            last_idx[name] = idx
        for entry in cup.get('ghosts', []):
            last_idx[entry[1]] = idx

    def dec(rating, name):
        missed = total_n - 1 - last_idx.get(name, 0)
        if missed <= GRACE: return round(rating, 1)
        return round(1500 + (rating - 1500) * (DECAY ** (missed - GRACE)), 1)

    out = []
    for name in rat:
        has_pod = sum(pods_d[name]) > 0
        if gp_d[name] < min_cups and not has_pod: continue
        raw = round(rat[name], 1)
        act = dec(rat[name], name)
        avg = round(total_pos_d[name] / avg_cups_d[name], 1) if avg_cups_d[name] > 0 else 0
        if peak_d is not None:
            pk = peak_d.get(name, raw)
        else:
            pk = max(h['rating'] for h in hist[name]) if hist[name] else raw
        h_list = [{'c': cup_num(h['cup']), 'r': h['rating'], 'p': h['position']}
                  for h in hist[name]]
        entry = {'n': name, 'a': act, 'r': raw}
        if unc_d is not None:
            entry['u'] = unc_d.get(name, 0)
        entry.update({
            'c': gp_d[name], 'b': best_d[name] if best_d[name] < 999 else 0,
            'v': avg, 'w': wins_d[name],
            'g': pods_d[name][0], 's': pods_d[name][1], 'z': pods_d[name][2],
            'p': round(pk, 1), 'h': h_list
        })
        out.append(entry)
    out.sort(key=lambda p: p['a'], reverse=True)
    return out

# ── Compute ───────────────────────────────────────────────────────────────

print("\nComputing TrueSkill (all cups)...")
ts_full = compute_trueskill(all_cups)
print("Computing TrueSkill (pure cups)...")
ts_pure = compute_trueskill(pure_cups)

print("Computing Standard ELO (all cups)...")
std_full = compute_standard_elo(all_cups)
print("Computing Standard ELO (pure cups)...")
std_pure = compute_standard_elo(pure_cups)

# ── Build JSON ────────────────────────────────────────────────────────────

# Merge into alldata.json (elo_engine.py already wrote weighted/season/glicko2;
# we append the alt-rank rating systems here).
alldata_path = _p('alldata.json')
with open(alldata_path, encoding='utf-8') as f:
    alldata = json.load(f)
alldata['standard']       = build_all_list(std_full, all_cups, min_cups=1)
alldata['standard_pure']  = build_all_list(std_pure, pure_cups, min_cups=1)
alldata['trueskill']      = build_all_list(ts_full, all_cups)
alldata['trueskill_pure'] = build_all_list(ts_pure, pure_cups)

# Cup dates — needed by altrank.html rolling tab. Sourced from cups.json
# (built by build_cups.py with date+log-mtime fallback).
cups_path = _p('cups.json')
with open(cups_path, encoding='utf-8') as f:
    cups_data = json.load(f)
cup_date_map = {}
for c in cups_data:
    if c.get('date') and c.get('id'):
        # id is like "COTD 144" or "Troll COTD 4" — skip non-pure cups (rolling = pure only)
        cid = c['id']
        if cid.startswith('Troll ') or 'Roulette' in cid:
            continue
        m = re.search(r'(\d+)', cid)
        if m:
            cup_date_map[int(m.group(1))] = c['date']
alldata['cupDates'] = cup_date_map
tmp = alldata_path + '.tmp'
with open(tmp, 'w') as f:
    json.dump(alldata, f, separators=(',', ':'))
os.replace(tmp, alldata_path)

size_kb = os.path.getsize(alldata_path) / 1024
print(f"\nalldata.json updated with standard + trueskill keys ({size_kb:.0f} KB)")

# ── Append standard rising keys to rising.json ────────────────────────────
# elo_engine.py writes rising.json with weighted/weighted_pure; altrank.html
# reads standard/standard_pure for its rolling/standard modes, so we append.

rising_path = _p('rising.json')
with open(rising_path, encoding='utf-8') as f:
    rising_out = json.load(f)

lookback_6m = rising_out['lookback_cup']
lookback_3m = rising_out['lookback_3m']

std_list = build_site_list(std_full, all_cups)
std_pure_list = build_site_list(std_pure, pure_cups)
rising_out['standard']      = build_rising_combined(std_list, lookback_6m, lookback_3m)
rising_out['standard_pure'] = build_rising_combined(std_pure_list, lookback_6m, lookback_3m)

with open(rising_path, 'w', encoding='utf-8') as f:
    json.dump(rising_out, f, indent=2)
print("rising.json updated with standard + standard_pure keys")

# ── Snapshot ──────────────────────────────────────────────────────────────
# DISABLED: altrank_snapshot.json is now built by snapshot.py (pre-cup baseline
# for arrows). Writing it here would overwrite the baseline with post-cup data.
# If altrank_snapshot.json doesn't exist, snapshot.py builds it.

# ── Sanity check ──────────────────────────────────────────────────────────

for label, data in [('Standard ELO', alldata['standard']), ('TrueSkill', alldata['trueskill'])]:
    top5 = data[:5]
    print(f"\n{label} top 5:")
    for i, p in enumerate(top5):
        print(f"  {i+1}. {p['n']:20s}  {p['a']:.1f}  ({p['c']} cups, {p['w']} wins)")
