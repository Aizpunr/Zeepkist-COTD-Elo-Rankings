"""
new_cup.py — Automated cup processing from COTDTracker mod logs

Usage:
  python new_cup.py 153 "[RAW]MapperName" --map "Real Map Name"

Fully automatic: snapshot -> parse log -> xlsx + JSON backup -> cup_meta.json
(map/mapper/date for build_cups.py) -> all ELO data -> localhost preview.
Push stays manual: verify on http://localhost:8000 first.
"""
import re, os, sys, json, subprocess, datetime
import openpyxl

# Force UTF-8 stdout so printing unicode aliases (e.g. the 𝒱V𝑜o𝒾i𝒹d𝒱 void name)
# in the alias-drift report can't crash the pipeline when stdout is redirected
# to a file (Windows defaults to cp1252, which can't encode those glyphs).
sys.stdout.reconfigure(encoding='utf-8')

_dir = os.path.dirname(os.path.abspath(__file__))
_p = lambda f: os.path.join(_dir, f)

LOG_PATH = r"C:\Program Files (x86)\Steam\steamapps\common\Zeepkist\BepInEx\LogOutput.log"
LIVE_LOG_PATH = r"C:\Program Files (x86)\Steam\steamapps\common\Zeepkist\BepInEx\LiveLeaderboardLogger.log"
COLS_PER_CUP = 6  # 4 data columns + 2 blank spacer

# ── Parse arguments ──
def _usage():
    print('Usage: python new_cup.py <cup_number> <mapper_name> --map "<map_name>"')
    print('                         [--exclude name1,name2,...] [--date YYYY-MM-DD]')
    print('                         [--log path] [--livelog path]')
    print('Example: python new_cup.py 153 "[MMM]Victor" --map "COTD - Blue Blitz"')
    print('Example: python new_cup.py 153 "PlusMicron" --map "Farewell" --exclude justMaki')
    sys.exit(1)

if len(sys.argv) < 3:
    _usage()

cup_num = int(sys.argv[1])
mapper = sys.argv[2]
cup_id = f"COTD {cup_num}"

# Optional --exclude flag for playtesters or other non-cup players
extra_excluded = []
if '--exclude' in sys.argv:
    idx = sys.argv.index('--exclude')
    if idx + 1 < len(sys.argv):
        extra_excluded = [n.strip() for n in sys.argv[idx + 1].split(',') if n.strip()]

excluded = {mapper} | set(extra_excluded)

# Optional --log / --livelog overrides: process a cup from a saved log file
# (e.g. an attendee's LogOutput.log when you missed the cup) instead of the
# live BepInEx log. When --log is given without --livelog, the local live log
# is NOT used — it would be a stale, unrelated session — so the alias SID
# check is skipped rather than run against the wrong cup.
def _arg_after(flag):
    if flag in sys.argv:
        i = sys.argv.index(flag)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return None

def _resolve(p):
    return p if os.path.isabs(p) else os.path.join(_dir, p)

# Required --map: the real map name, written to cup_meta.json so build_cups.py
# needs no hand-edit (map_index / CUP_DATES additions are dead — see cup_meta).
map_name = _arg_after('--map')
if not map_name:
    print("ERROR: --map \"<map_name>\" is required (goes to cup_meta.json).")
    _usage()

# Optional --date YYYY-MM-DD override. Without it the cup date is derived from
# the log file's mtime; a log written before 06:00 counts as the PREVIOUS day
# (cups run Saturday evening and often end past midnight — COTD 148 gotcha).
_date_override = _arg_after('--date')
if _date_override and not re.fullmatch(r'\d{4}-\d{2}-\d{2}', _date_override):
    print(f"ERROR: --date must be YYYY-MM-DD, got {_date_override!r}")
    sys.exit(1)

_log_override = _arg_after('--log')
_livelog_override = _arg_after('--livelog')
if _log_override:
    LOG_PATH = _resolve(_log_override)
    LIVE_LOG_PATH = _resolve(_livelog_override) if _livelog_override else None
    print(f"Log override: {LOG_PATH}")
    print(f"Livelog override: {LIVE_LOG_PATH if LIVE_LOG_PATH else '(none — alias SID check skipped)'}")
elif _livelog_override:
    LIVE_LOG_PATH = _resolve(_livelog_override)

print(f"Processing {cup_id}, mapper: {mapper}, map: {map_name}")
if extra_excluded:
    print(f"Also excluding: {', '.join(extra_excluded)}")
print()

# ── 0a. Idempotency gate + auto-detect xlsx file/columns ──
# Refuse to process a cup that's already in the xlsx or cup_meta.json: a
# re-run would append a DUPLICATE column and corrupt every downstream build.
# (The xlsx is gitignored — recovery is backups/<xlsx>.pre_cup_<N>.xlsx, not git.)
elo_py_path = _p('elo_engine.py')
with open(elo_py_path, encoding='utf-8') as f:
    elo_src = f.read()

xlsx_matches = re.findall(r"parse_file\(_p\('(COTD \d+-\d+\.xlsx)'\)\)", elo_src)
if not xlsx_matches:
    print("ERROR: Could not find COTD xlsx reference in elo_engine.py")
    sys.exit(1)
current_xlsx = xlsx_matches[-1]
xlsx_path = _p(current_xlsx)

print(f"Current xlsx: {current_xlsx}")

wb = openpyxl.load_workbook(xlsx_path)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

existing_cup_ids = {str(v).strip() for v in rows[1] if v}  # row 2 = cup headers
meta_path = _p('cup_meta.json')
cup_meta = {}
if os.path.exists(meta_path):
    with open(meta_path, encoding='utf-8') as f:
        cup_meta = json.load(f)

if cup_id in existing_cup_ids or cup_id in cup_meta:
    where = []
    if cup_id in existing_cup_ids:
        where.append(current_xlsx)
    if cup_id in cup_meta:
        where.append('cup_meta.json')
    print(f"ERROR: {cup_id} is already processed (found in {' + '.join(where)}).")
    print(f"  To reprocess: restore backups/{current_xlsx}.pre_cup_{cup_num}.xlsx")
    print(f"  over {current_xlsx}, remove the {cup_id!r} entry from cup_meta.json,")
    print(f"  then re-run. Otherwise check the cup number.")
    sys.exit(1)

# Find rightmost Position header to determine next column
max_pos_col = 0
for ci, val in enumerate(rows[4]):  # Row 5 (0-indexed row 4) has headers
    if val == 'Position':
        max_pos_col = ci + 1  # Convert to 1-based
col_start = max_pos_col + COLS_PER_CUP
print(f"Last Position header at column {max_pos_col}, new cup at column {col_start}")

# ── 0b. Snapshot previous standings (was a separate manual step) ──
# Explicit target = previous cup, so a partial re-run can never snapshot the
# post-update standings by accident. snapshot.py backs up the old snapshot
# itself and fails loud if alldata.json is half-written.
print("=" * 50)
print(f"Running snapshot.py {cup_num - 1}...")
print("=" * 50)
r = subprocess.run([sys.executable, _p('snapshot.py'), str(cup_num - 1)], cwd=_dir)
if r.returncode != 0:
    print("ERROR: snapshot.py failed — fix alldata.json before processing the cup.")
    sys.exit(1)
print()

# ── 1. Save raw log + parse mod logs ──
log_dir = _p('cup logs')
os.makedirs(log_dir, exist_ok=True)
import shutil
log_backup = os.path.join(log_dir, f'cotd_{cup_num}.log')
if os.path.abspath(LOG_PATH) != os.path.abspath(log_backup):
    shutil.copy2(LOG_PATH, log_backup)
    print(f"Raw log saved: {log_backup}")
else:
    print(f"Raw log already in place: {log_backup}")
live_log_backup = os.path.join(log_dir, f'cotd_{cup_num}_liveleaderboard.log')
if LIVE_LOG_PATH and os.path.exists(LIVE_LOG_PATH):
    shutil.copy2(LIVE_LOG_PATH, live_log_backup)
    print(f"Live log saved: {live_log_backup}")
else:
    print(f"⚠ Live log not found at {LIVE_LOG_PATH} — alias SID check will be skipped")
    live_log_backup = None

with open(LOG_PATH, encoding='utf-8', errors='replace') as f:
    lines = [l for l in f.readlines() if 'COTDTracker' in l]

if not lines:
    print("ERROR: No COTDTracker lines found in log file.")
    sys.exit(1)

rounds = []
current_round = []
for line in lines:
    if 'Doing eliminations with leaderboard' in line:
        if current_round:
            rounds.append(current_round)
        current_round = []
    elif 'Eliminating ' in line or 'Player ' in line:
        current_round.append(line)
if current_round:
    rounds.append(current_round)

if not rounds:
    print("ERROR: No elimination rounds found in log.")
    sys.exit(1)

# Build elimination order
# For each eliminated player, track:
#   display_time: their time in their elim round (or 'DNF' if they DNF'd it)
#   dnf:          True if they DNF'd their elimination round (determines tie)
#
# NOTE: display_time is ONLY from the elim round's own log entries. We used to
# fall back to last_known_time (some prior round's time) for DNFs, but that was
# misleading — it showed a time from a different round next to the round's
# elim number. DNFs now stay DNF.
elim_order = []
actual_round = 0
for rnd in rounds:
    player_times = {}
    eliminated_names = []
    for line in rnd:
        m = re.search(r'Player (.+?): Time: (.+)', line)
        if m:
            name = m.group(1).strip()
            time_str = m.group(2).strip()
            player_times[name] = time_str
        m2 = re.search(r'Eliminating (?:DNF|on time): (.+)', line)
        if m2:
            name = m2.group(1).strip()
            if name not in eliminated_names:
                eliminated_names.append(name)
    if not eliminated_names:
        continue
    actual_round += 1
    for name in eliminated_names:
        if name not in excluded:
            elim_round_time = player_times.get(name, 'DNF')
            dnf = (elim_round_time == 'DNF')
            elim_order.append((name, elim_round_time, actual_round, dnf))

# Find winner
all_named = set()
for rnd in rounds:
    for line in rnd:
        m = re.search(r'Player (.+?): Time:', line)
        if m:
            all_named.add(m.group(1).strip())
# Excluded names that never appear in the log did not exclude anyone: either
# they genuinely didn't play, or the raw name is wrong (COTD 152: mapper passed
# as "Victor" while the log had "[MMM]Victor" — he got counted as a player).
for name in sorted(excluded):
    if name not in all_named:
        print(f"⚠ WARNING: excluded name {name!r} not found in the log — "
              f"either they didn't play, or this isn't their exact raw in-game name.")

elim_set = {e[0] for e in elim_order}
winners = [n for n in all_named if n not in elim_set and n not in excluded]
if not winners:
    print("ERROR: Could not determine winner.")
    sys.exit(1)
if len(winners) > 1:
    # "Named but never eliminated" resolved to 2+ players — usually a
    # mid-cup disconnect the tracker never eliminated, or a truncated log.
    # An unattended run must not guess; add the non-winner(s) to --exclude
    # or fix the log and re-run.
    print(f"ERROR: winner is ambiguous — {len(winners)} players were never eliminated:")
    for n in winners:
        print(f"  - {n!r}")
    print("Pick the real winner from the VOD/log; re-run with the others handled")
    print("(e.g. --exclude for non-players, or use a more complete log).")
    sys.exit(1)
winner = winners[0]

winner_time = None
for line in reversed(lines):
    m = re.search(r'Player ' + re.escape(winner) + r': Time: (.+)', line)
    if m:
        winner_time = m.group(1).strip()
        break

# Build leaderboard: within each elimination round, finishers get distinct
# positions ordered by their elim-round time (faster = better), then DNFs
# tie at the bottom of that round.
elim_order.reverse()
leaderboard = [(winner, winner_time, None, 1)]
pos = 2
i = 0
while i < len(elim_order):
    rnd = elim_order[i][2]
    group = []
    while i < len(elim_order) and elim_order[i][2] == rnd:
        group.append(elim_order[i])
        i += 1
    # Split by dnf flag — finishers get distinct positions by elim-round time,
    # DNFs (no valid time in elim round) all tie at the bottom of this round.
    finishers = []
    dnfs = []
    for name, display_time, r, dnf in group:
        if dnf:
            dnfs.append((name, display_time, r))
        else:
            try:
                t = float(str(display_time).replace(',', '.'))
                finishers.append((name, display_time, r, t))
            except (ValueError, TypeError):
                dnfs.append((name, display_time, r))
    finishers.sort(key=lambda x: x[3])
    for name, display_time, r, _ in finishers:
        leaderboard.append((name, display_time, r, pos))
        pos += 1
    if dnfs:
        dnf_pos = pos
        for name, display_time, r in dnfs:
            leaderboard.append((name, display_time, r, dnf_pos))
        pos += len(dnfs)

print(f"Parsed {len(leaderboard)} players (excluded: {', '.join(sorted(excluded))})")
print(f"Winner: {winner} ({winner_time})")
print(f"Rounds in log: {len(rounds)}")
print()

# Find fastest time. Round numbering matches the elim_order loop above:
# the first leaderboard with no eliminations is the discovery/warmup round (R0),
# Round 1 is the first round that actually eliminates someone.
fastest_time = None
fastest_name = None
fastest_round = None
actual_round = 0
for rnd in rounds:
    has_elim = any(re.search(r'Eliminating (?:DNF|on time):', line) for line in rnd)
    if has_elim:
        actual_round += 1
    rnd_num = actual_round if has_elim else 0
    for line in rnd:
        m = re.search(r'Player (.+?): Time: (.+)', line)
        if m:
            name = m.group(1).strip()
            time_str = m.group(2).strip()
            if time_str != 'DNF':
                t = float(time_str.replace(',', '.'))
                if fastest_time is None or t < fastest_time:
                    fastest_time = t
                    fastest_name = name
                    fastest_round = rnd_num

# ── 1b. Cross-check leaderboard names against livelog Steam IDs ──
# Catches alias drift: a player who joins a new clan and ends up tracked as a
# fresh identity (e.g. "[NewB]Zeus" → "[SLOW]Zeus" → "Zeus"). Same Steam ID =
# same player; mismatches between the lobby name and the canonical name in
# steam_ids.json get flagged with a copy-pasteable CANONICAL entry suggestion.
def check_aliases_against_livelog(leaderboard, live_log_path, steam_ids_path, canonical_src_path,
                                  report_path=None):
    if not live_log_path or not os.path.exists(live_log_path):
        print("⚠ ALIAS CHECK SKIPPED — no live log to read")
        return
    if not os.path.exists(steam_ids_path):
        print(f"⚠ ALIAS CHECK SKIPPED — no steam_ids.json at {steam_ids_path}")
        return
    # name -> SID from ROSTER lines
    roster = {}
    with open(live_log_path, encoding='utf-8-sig', errors='replace') as f:
        for line in f:
            mr = re.search(r'ROSTER\|(\d+)\|([^|]*)\|([^|]*)', line)
            if mr:
                sid = mr.group(1)
                for n in (mr.group(2).strip(), mr.group(3).strip()):
                    if n:
                        roster[n] = sid
    # SID -> canonical from steam_ids.json
    with open(steam_ids_path, encoding='utf-8') as f:
        sids = json.load(f)
    sid_to_canon = {v: k for k, v in sids.items()}
    # alias -> canonical from CANONICAL block in elo_engine.py
    with open(canonical_src_path, encoding='utf-8') as f:
        src = f.read()
    canon_block = re.search(r'CANONICAL\s*=\s*\{(.+?)^\}', src, re.MULTILINE | re.DOTALL)
    alias_to_canon, canon_names = {}, set()
    if canon_block:
        canon_dict = eval('{' + canon_block.group(1) + '}')
        canon_names = set(canon_dict.keys())
        for cn, aliases in canon_dict.items():
            for a in aliases:
                alias_to_canon[a] = cn

    drifts, new_players = [], []
    for entry in leaderboard:
        name = entry[0]
        sid = roster.get(name)
        if not sid:
            bare = re.sub(r'[\[\{].*?[\]\}]\s*', '', name).strip()
            sid = roster.get(bare)
        if not sid:
            continue
        canon_for_sid = sid_to_canon.get(sid)
        if canon_for_sid is None:
            new_players.append((name, sid))
            continue
        if name in alias_to_canon:
            resolved = alias_to_canon[name]
        elif name in canon_names or name == canon_for_sid:
            resolved = name if name in canon_names else canon_for_sid
        else:
            resolved = name  # would be tracked as the raw name = NOT merged
        if resolved != canon_for_sid:
            drifts.append((name, sid, canon_for_sid))

    if not drifts and not new_players:
        print("Alias check: all names match steam_ids.json canonicals ✓")
        return
    # Build the report as lines: printed to console AND written to a file so
    # an unattended run's findings aren't lost when the console closes. The
    # report never blocks the pipeline — until CANONICAL/steam_ids.json are
    # curated, the player is simply tracked under their raw name.
    out = []
    out.append("=" * 50)
    out.append("ALIAS CHECK — livelog SIDs vs steam_ids.json")
    out.append("=" * 50)
    if drifts:
        out.append(f"\n{len(drifts)} alias drift(s):")
        for name, sid, canon in drifts:
            out.append(f"  - {name!r} (sid {sid}) should map to canonical {canon!r}")
        out.append("\nUpdate CANONICAL in elo_engine.py — append these to the matching entry:")
        from collections import defaultdict
        by_canon = defaultdict(list)
        for name, _, canon in drifts:
            by_canon[canon].append(name)
        for canon, names in by_canon.items():
            names_lit = ', '.join(repr(n) for n in sorted(set(names)))
            out.append(f"  '{canon}': [..., {names_lit}],")
    if new_players:
        out.append(f"\n{len(new_players)} new player(s) — no entry in steam_ids.json:")
        for name, sid in new_players:
            out.append(f"  - {name!r} (sid {sid})")
        out.append("\nIf any are returning under a new name, add to steam_ids.json:")
        for name, sid in new_players:
            out.append(f'  "{name}": "{sid}",')
    print()
    print('\n'.join(out))
    print()
    if report_path:
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(out) + '\n')
        print(f"Alias report saved: {os.path.relpath(report_path, _dir)}")

check_aliases_against_livelog(
    leaderboard,
    live_log_backup,
    _p('steam_ids.json'),
    _p('elo_engine.py'),
    report_path=_p(os.path.join('cup logs', f'cotd_{cup_num}_alias_report.txt')),
)

# ── 3. Write to xlsx (file/columns detected in step 0a) ──
# Pre-write backup: the xlsx is gitignored, so this copy is the ONLY way to
# recover from a bad/duplicate write (see the idempotency gate's message).
os.makedirs(_p('backups'), exist_ok=True)
_xlsx_backup = _p(os.path.join('backups', f'{current_xlsx}.pre_cup_{cup_num}.xlsx'))
shutil.copy2(xlsx_path, _xlsx_backup)
print(f"xlsx backup: backups/{os.path.basename(_xlsx_backup)}")

ws.cell(row=2, column=col_start, value=cup_id)
ws.cell(row=3, column=col_start, value=f'Map: {map_name} by {mapper}')

if fastest_time is not None:
    ws.cell(row=4, column=col_start + 2,
            value=f'Fastest Time: {fastest_time:.3f} by {fastest_name} in Round {fastest_round}')

ws.cell(row=5, column=col_start, value='Position')
ws.cell(row=5, column=col_start + 1, value='Name')
ws.cell(row=5, column=col_start + 2, value='Elim Time')
ws.cell(row=5, column=col_start + 3, value='Elim Round')

for i, (name, time, rnd, position) in enumerate(leaderboard):
    row = 6 + i
    ws.cell(row=row, column=col_start, value=position)
    ws.cell(row=row, column=col_start + 1, value=name)
    if time == 'DNF':
        ws.cell(row=row, column=col_start + 2, value='DNF')
    else:
        t = float(time.replace(',', '.'))
        ws.cell(row=row, column=col_start + 2, value=round(t * 1000))
    if rnd is not None:
        ws.cell(row=row, column=col_start + 3, value=rnd)

def _recycle(path):
    """Send a file to the Windows Recycle Bin — never permanently delete."""
    r = subprocess.run([
        'powershell', '-NoProfile', '-Command',
        'Add-Type -AssemblyName Microsoft.VisualBasic; '
        "[Microsoft.VisualBasic.FileIO.FileSystem]::DeleteFile("
        f"'{path}', 'OnlyErrorDialogs', 'SendToRecycleBin')"
    ], capture_output=True, text=True)
    return r.returncode == 0

# Rename xlsx if needed
m = re.match(r'COTD (\d+)-(\d+)\.xlsx', current_xlsx)
if m:
    start_num, end_num = int(m.group(1)), int(m.group(2))
    if cup_num > end_num:
        new_xlsx = f'COTD {start_num}-{cup_num}.xlsx'
        new_path = _p(new_xlsx)
        wb.save(new_path)
        print(f"Saved as {new_xlsx} (renamed from {current_xlsx})")

        # Point elo_engine.py at the new file BEFORE touching the old one, so
        # a crash in between never leaves the engine referencing a missing file.
        new_src = elo_src.replace(f"'{current_xlsx}'", f"'{new_xlsx}'")
        with open(elo_py_path, 'w', encoding='utf-8') as f:
            f.write(new_src)
        print(f"Updated elo_engine.py: {current_xlsx} -> {new_xlsx}")

        # The old file is the master results spreadsheet — Recycle Bin only,
        # never os.remove. If recycling fails, leave it in place (harmless:
        # every script resolves the xlsx via elo_engine.py source, not glob).
        if new_path != xlsx_path and os.path.exists(xlsx_path):
            if _recycle(xlsx_path):
                print(f"Old xlsx sent to Recycle Bin: {current_xlsx}")
            else:
                print(f"WARNING: could not recycle {current_xlsx} — left in place, remove manually")
    else:
        wb.save(xlsx_path)
        print(f"Saved to {current_xlsx}")
else:
    wb.save(xlsx_path)
    print(f"Saved to {current_xlsx}")

# ── 4. Write JSON backup ──
cup_json = {
    'cup': cup_id,
    'cup_num': cup_num,
    'mapper': mapper,
    'players': [
        {'pos': position, 'name': name, 'time': time, 'round': rnd}
        for name, time, rnd, position in leaderboard
    ]
}
json_path = _p(f'cup_{cup_num}.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(cup_json, f, ensure_ascii=False, indent=2)
print(f"JSON backup: cup_{cup_num}.json")
print()

# ── 4b. Write cup_meta.json (map/mapper/date for build_cups.py) ──
# Replaces the manual map_index + CUP_DATES edits in build_cups.py source.
# Date: --date override, else the log's mtime — but a log written before
# 06:00 belongs to the PREVIOUS day (cups end past midnight; COTD 148's log
# was stamped early Sunday 06-14 for the Saturday 06-13 cup).
if _date_override:
    cup_date = _date_override
else:
    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(LOG_PATH))
    d = mtime.date()
    if mtime.hour < 6:
        d -= datetime.timedelta(days=1)
    cup_date = d.isoformat()

cup_meta[cup_id] = {'map': map_name, 'mapper': mapper, 'date': cup_date}
tmp = meta_path + '.tmp'
with open(tmp, 'w', encoding='utf-8') as f:
    json.dump(cup_meta, f, ensure_ascii=False, indent=2)
    f.write('\n')
os.replace(tmp, meta_path)
print(f"cup_meta.json: {cup_id} = {map_name!r} by {mapper!r}, {cup_date}")
print()

# ── 5. Run elo_engine.py ──
print("=" * 50)
print("Running elo_engine.py...")
print("=" * 50)
result = subprocess.run([sys.executable, _p('elo_engine.py')], cwd=_dir)
if result.returncode != 0:
    print("ERROR: elo_engine.py failed!")
    sys.exit(1)
print()

# ── 6. Run build_cups.py ──
print("=" * 50)
print("Running build_cups.py...")
print("=" * 50)
result = subprocess.run([sys.executable, _p('build_cups.py')], cwd=_dir)
if result.returncode != 0:
    print("ERROR: build_cups.py failed!")
    sys.exit(1)
print()

# ── 7. Rebuild Cool Stats ──
# build_altrank.py is NOT just a cool stat: it's the second writer of
# alldata.json (elo_engine.py overwrites the file with 5 keys, build_altrank
# re-appends standard/trueskill/cupDates). If it fails, alldata.json stays
# half-written, altrank.html breaks and the next snapshot.py run crashes —
# this shipped to the live site once (2026-06-02) and blanked the leaderboard.
# So its failure aborts the pipeline; the rest only warn.
cool_stats = [
    ('build_big3.py', 'Big 3 H2H'),
    ('build_giantkillers.py', 'Giant Killers'),
    ('build_consistency.py', 'Consistency Index'),
    ('elo_stability.py', 'ELO Stability'),
    ('build_fastest.py', 'Fastest Times'),
    ('build_whatif.py', 'What-If ELO'),
    ('build_streaks.py', 'Streak Records'),
    ('build_altrank.py', 'Alt Rankings'),
]
FATAL_STATS = {'build_altrank.py'}
failed_steps = []
print("=" * 50)
print("Rebuilding Cool Stats...")
print("=" * 50)
for script, label in cool_stats:
    path = _p(script)
    if os.path.exists(path):
        r = subprocess.run([sys.executable, path], cwd=_dir)
        if r.returncode != 0:
            if script in FATAL_STATS:
                print(f"ERROR: {label} ({script}) failed — alldata.json is "
                      f"HALF-WRITTEN (missing standard/trueskill/cupDates).")
                print(f"Fix the error and re-run: python {script}")
                print("Do NOT push until it succeeds.")
                sys.exit(1)
            failed_steps.append(f"{label} ({script})")
            print(f"  WARNING: {label} ({script}) failed!")
        else:
            print(f"  {label} OK")
    else:
        if script in FATAL_STATS:
            print(f"ERROR: {script} not found — alldata.json would stay half-written.")
            sys.exit(1)
        failed_steps.append(f"{label} ({script}) — not found")
        print(f"  SKIP: {script} not found")
print()

# ── 8. Refresh cross-comp ranking + SOF mod data ──
# Cross-comp pipeline: rebuilds allsofdata.json + allcompdata.json + the
# SOF mod's elo_pool.json (which now uses cross-comp ELO as its primary
# source, GTR rank regression as fallback). Replaces the old COTD-only
# join_cotd_gtr.py path.
print("=" * 50)
print("Refreshing cross-comp ranking + SOF data...")
print("=" * 50)
crosscomp_script = r"C:\Users\rafa\Desktop\Claude\zeepkist holistic\refresh.py"
sof_repo_pool = r"C:\Users\rafa\Desktop\Claude\zeepkist mod\Zeepkist-Strength-of-Field\elo_pool.json"

sof_ok = False
if os.path.exists(crosscomp_script):
    r = subprocess.run([sys.executable, crosscomp_script],
                       cwd=os.path.dirname(crosscomp_script))
    if r.returncode == 0:
        print(f"  Cross-comp + SOF elo_pool.json refreshed.")
        sof_ok = True
    else:
        print(f"  WARNING: cross-comp refresh failed (returncode={r.returncode}).")
        print(f"  Run manually: python \"{crosscomp_script}\"")
else:
    print(f"  SKIP: refresh script not found at {crosscomp_script}")
print()

# ── 8b. Refresh the COTD-only SOF pool (the mod's `!sof cup` data) ──
# elo_pool_cotd.json is the historical / COTD-only counterpart to the
# cross-comp elo_pool.json above. It only changes when COTD ratings change,
# so it lives here in the COTD pipeline. Non-fatal: a GTR outage (it hits
# graphql.zeepki.st) shouldn't break cup processing.
print("=" * 50)
print("Refreshing COTD SOF pool (mod's !sof cup data)...")
print("=" * 50)
cotd_pool_script = r"C:\Users\rafa\Desktop\Claude\zeepkist mod\gtr_analysis\join_cotd_gtr.py"
sof_repo_cotd_pool = r"C:\Users\rafa\Desktop\Claude\zeepkist mod\Zeepkist-Strength-of-Field\elo_pool_cotd.json"

cotd_pool_ok = False
if os.path.exists(cotd_pool_script):
    r = subprocess.run([sys.executable, cotd_pool_script],
                       cwd=os.path.dirname(cotd_pool_script))
    if r.returncode == 0:
        print(f"  COTD elo_pool_cotd.json refreshed.")
        cotd_pool_ok = True
    else:
        print(f"  WARNING: COTD pool refresh failed (returncode={r.returncode}).")
        print(f"  Run manually: python \"{cotd_pool_script}\"")
else:
    print(f"  SKIP: COTD pool script not found at {cotd_pool_script}")
print()

# ── 9. Summary ──
print("=" * 50)
print(f"{cup_id} COMPLETE")
if failed_steps:
    print(f"  !! {len(failed_steps)} non-fatal step(s) FAILED — fix before pushing:")
    for s in failed_steps:
        print(f"     - {s}")
print(f"  Players: {len(leaderboard)}")
print(f"  Winner: {winner}")
if fastest_time:
    print(f"  Fastest: {fastest_time:.3f} by {fastest_name}")
print(f"  Columns: {col_start}-{col_start+3}")
if sof_ok or cotd_pool_ok:
    print(f"  SOF data: refreshed (commit+push the SOF repo too)")
print()
print("Next steps (after verifying on localhost):")
print(f"  - Git commit + push (COTD repo) — REMEMBER to stage cup_meta.json AND cup_{cup_num}.json (lexertools last-cup view fetches it)")
if sof_ok or cotd_pool_ok:
    print(f"  - Git commit + push (SOF repo) — stage all three:")
    if sof_ok:
        print(f"      {sof_repo_pool}")
    if cotd_pool_ok:
        print(f"      {sof_repo_cotd_pool}")
    print(f"      docs/allcompdata.json (cross-comp site data — easy to miss)")
print("=" * 50)

# ── 10. Localhost preview ──
# The run ends with the updated site visible on localhost:8000 (COTD's
# reserved port) — verify there BEFORE pushing. If nothing is listening,
# start a detached static server that outlives this script.
import socket
_serving = False
try:
    with socket.create_connection(('127.0.0.1', 8000), timeout=0.3):
        _serving = True
except OSError:
    pass
if _serving:
    print("localhost:8000 already serving — updated data is live there.")
else:
    _flags = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP
    subprocess.Popen([sys.executable, '-m', 'http.server', '8000'],
                     cwd=_dir, creationflags=_flags,
                     stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    print("Started http.server on localhost:8000 (detached).")
try:
    os.startfile('http://localhost:8000')
    print("Opened http://localhost:8000 — verify the cup, then push manually.")
except OSError:
    print("Open http://localhost:8000 to verify the cup, then push manually.")
