import re
import openpyxl

log_path = r"C:\Program Files (x86)\Steam\steamapps\common\Zeepkist\BepInEx\LogOutput.log"
with open(log_path) as f:
    lines = [l for l in f.readlines() if 'COTDTracker' in l]

mapper = '[20x]K410K3N'

# Parse rounds
rounds = []
current_round = []
for line in lines:
    if 'Doing eliminations with leaderboard' in line:
        if current_round:
            rounds.append(current_round)
        current_round = []
    elif 'Eliminating ' in line or 'Player ' in line:
        current_round.append(line)
if current_round:
    rounds.append(current_round)

# Build elimination order
elim_order = []
for round_num, rnd in enumerate(rounds, 1):
    player_times = {}
    eliminated_names = []
    for line in rnd:
        m = re.search(r'Player (.+?): Time: (.+)', line)
        if m:
            player_times[m.group(1).strip()] = m.group(2).strip()
        m2 = re.search(r'Eliminating (?:DNF|on time): (.+)', line)
        if m2:
            name = m2.group(1).strip()
            if name not in eliminated_names:
                eliminated_names.append(name)
    for name in eliminated_names:
        if name != mapper:
            time = player_times.get(name, 'DNF')
            elim_order.append((name, time, round_num))

# Find winner
all_named = set()
for rnd in rounds:
    for line in rnd:
        m = re.search(r'Player (.+?): Time:', line)
        if m:
            all_named.add(m.group(1).strip())
elim_set = {e[0] for e in elim_order}
winner = [n for n in all_named if n not in elim_set and n != mapper][0]

winner_time = None
for line in reversed(lines):
    m = re.search(r'Player ' + re.escape(winner) + r': Time: (.+)', line)
    if m:
        winner_time = m.group(1).strip()
        break

# Build leaderboard
elim_order.reverse()
leaderboard = [(winner, winner_time, None)] + [(n, t, r) for n, t, r in elim_order]

# Write to spreadsheet
# Cup 134 goes in columns 19-22 (Position, Name, Elim Time, Elim Round)
xlsx_path = 'COTD 131-133.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['COTD 131-135']

# Header rows
col_start = 19
ws.cell(row=2, column=col_start, value='COTD 134')
ws.cell(row=3, column=col_start, value='Map: COTD - Urbs Noctu by [20x]K410K3N')

# Find fastest time
fastest_time = None
fastest_name = None
fastest_round = None
for rnd_num, rnd in enumerate(rounds, 1):
    for line in rnd:
        m = re.search(r'Player (.+?): Time: (.+)', line)
        if m:
            name = m.group(1).strip()
            time_str = m.group(2).strip()
            if time_str != 'DNF':
                t = float(time_str.replace(',', '.'))
                if fastest_time is None or t < fastest_time:
                    fastest_time = t
                    fastest_name = name
                    fastest_round = rnd_num

ws.cell(row=4, column=col_start + 2, value=f'Fastest Time: {fastest_time:.3f} by {fastest_name} in Round {fastest_round}')

# Column headers
ws.cell(row=5, column=col_start, value='Position')
ws.cell(row=5, column=col_start + 1, value='Name')
ws.cell(row=5, column=col_start + 2, value='Elim Time')
ws.cell(row=5, column=col_start + 3, value='Elim Round')

# Data
for i, (name, time, rnd) in enumerate(leaderboard):
    row = 6 + i
    ws.cell(row=row, column=col_start, value=i + 1)
    ws.cell(row=row, column=col_start + 1, value=name)
    if time == 'DNF':
        ws.cell(row=row, column=col_start + 2, value='DNF')
    else:
        # Store as number (milliseconds style like Lexer does)
        t = float(time.replace(',', '.'))
        ws.cell(row=row, column=col_start + 2, value=round(t * 1000))
    if rnd is not None:
        ws.cell(row=row, column=col_start + 3, value=rnd)

wb.save(xlsx_path)
print(f'Cup 134 written to {xlsx_path}, columns {col_start}-{col_start+3}, {len(leaderboard)} players')
