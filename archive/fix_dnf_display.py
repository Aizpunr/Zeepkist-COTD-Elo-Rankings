"""
fix_dnf_display.py — one-off patch to set DNF rows in xlsx to 'DNF'.

Reads each cup log from `cup logs/`, determines which players truly DNF'd
their elimination round, and overwrites the xlsx time cell for those rows
to literal 'DNF' (instead of the fallback time from a prior round).

ELO calculations use position only, not display time, so this is purely cosmetic.
Run once to clean up legacy cups that were processed before the new_cup.py fix.
"""
import openpyxl, re, os, sys
sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = 'COTD 131-140.xlsx'
CUPS = [136, 137, 139, 140]

def parse_log(path):
    """Per-elim-round DNF detection. Returns dict name -> True/False (is DNF)."""
    with open(path, encoding='utf-8', errors='replace') as f:
        lines = [l for l in f.readlines() if 'COTDTracker' in l]
    rounds = []
    current = []
    for line in lines:
        if 'Doing eliminations with leaderboard' in line:
            if current: rounds.append(current)
            current = []
        elif 'Eliminating ' in line or 'Player ' in line:
            current.append(line)
    if current: rounds.append(current)
    result = {}
    for rnd in rounds:
        pt = {}
        elim = []
        for line in rnd:
            m = re.search(r'Player (.+?): Time: (.+)', line)
            if m: pt[m.group(1).strip()] = m.group(2).strip()
            m2 = re.search(r'Eliminating (?:DNF|on time): (.+)', line)
            if m2:
                name = m2.group(1).strip()
                if name not in elim: elim.append(name)
        for name in elim:
            if name in result: continue
            result[name] = (pt.get(name, 'DNF') == 'DNF')
    return result

def strip_tag(s):
    return re.sub(r'^\[.*?\]\s*', '', s).strip()

wb = openpyxl.load_workbook(XLSX_PATH)
patches = []

for cup in CUPS:
    log_path = os.path.join('cup logs', f'cotd_{cup}.log')
    if not os.path.exists(log_path):
        print(f'COTD {cup}: log missing, skip')
        continue
    log = parse_log(log_path)
    dnf_set = {name for name, is_dnf in log.items() if is_dnf}
    dnf_set_stripped = {strip_tag(n) for n in dnf_set}

    target = f'COTD {cup}'
    found_loc = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == 'Position':
                    # Check header above
                    pos_row, pos_col = cell.row, cell.column
                    for sr in range(pos_row - 1, max(pos_row - 5, 0), -1):
                        hv = ws.cell(sr, pos_col).value
                        if hv and str(hv).strip() == target:
                            found_loc = (sheet_name, pos_row, pos_col)
                            break
                    if found_loc: break
            if found_loc: break
        if found_loc: break

    if not found_loc:
        print(f'COTD {cup}: header not found in xlsx, skip')
        continue

    sheet_name, pos_row, pos_col = found_loc
    ws = wb[sheet_name]
    cup_patches = []
    # Scan rows below header
    for r in range(pos_row + 1, ws.max_row + 1):
        name_cell = ws.cell(r, pos_col + 1)
        time_cell = ws.cell(r, pos_col + 2)
        if name_cell.value is None: break
        name_str = str(name_cell.value).strip()
        if name_str.startswith('*'): continue
        stripped = strip_tag(name_str)
        is_dnf = (name_str in dnf_set) or (stripped in dnf_set_stripped)
        if is_dnf and time_cell.value != 'DNF':
            cup_patches.append((r, pos_col + 2, name_str, time_cell.value))
            time_cell.value = 'DNF'

    print(f'COTD {cup} ({sheet_name}): {len(cup_patches)} rows patched')
    for r, c, n, old in cup_patches[:5]:
        print(f'  row {r} ({n}): {old} -> DNF')
    if len(cup_patches) > 5:
        print(f'  ... and {len(cup_patches) - 5} more')
    patches.extend(cup_patches)

if patches:
    wb.save(XLSX_PATH)
    print(f'\nSaved {XLSX_PATH} with {len(patches)} DNF patches.')
else:
    print('\nNo patches applied.')
