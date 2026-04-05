"""
build_fastest.py — extract fastest times from all COTD xlsx spreadsheets.
Outputs fastest.json sorted by time ascending.
"""
import json, os, re, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base = os.path.dirname(os.path.abspath(__file__))
def _p(f): return os.path.join(base, f)

# Import CANONICAL aliases from elo_75.py
def load_aliases():
    """Parse CANONICAL dict from elo_75.py source."""
    name_map = {}
    lines = open(_p('elo_75.py'), encoding='utf-8').readlines()
    collecting = False
    buf = []
    for line in lines:
        if not collecting and re.match(r'^CANONICAL\s*=\s*\{', line):
            collecting = True
        if collecting:
            buf.append(line)
            # Stop when we hit a line that's just "}" (the closing of CANONICAL)
            if line.strip() == '}':
                break
    if not buf:
        return name_map
    block = ''.join(buf).split('=', 1)[1].strip()
    canonical = eval(block)
    for canon, aliases in canonical.items():
        for alias in aliases:
            name_map[alias] = canon
    return name_map

NAME_MAP = load_aliases()

def normalize_name(name):
    """Resolve tagged/aliased name to canonical."""
    if name in NAME_MAP:
        return NAME_MAP[name]
    # Strip tag and check
    stripped = re.sub(r'^\[.*?\]\s*', '', name).strip()
    if stripped in NAME_MAP:
        return NAME_MAP[stripped]
    return stripped

FILES = [
    'Zeepkist COTDs 1-25.xlsx',
    'Zeepkist COTDs 26-50.xlsx',
    'Zeepkist COTDs 51-75.xlsx',
    'COTDs 76-100.xlsx',
    'COTDs 101-125.xlsx',
    'COTD 126-130.xlsx',
    'COTD 131-138.xlsx',
    'cup roulette.xlsx',
    'Troll cup.xlsx',
]

# Regex for standard format: "Fastest Time: 45.823 by justMaki in Round 12"
# Also handles minutes format: "1:36.225 by ..."
RE_FT = re.compile(
    r'Fastest Time:\s*'
    r'(?:(\d+):)?'          # optional minutes
    r'([\d.]+)'             # seconds
    r'\s+by\s+'
    r'(.+?)'                # player name
    r'(?:\s+in\s+(.+))?$'  # optional round info
)


def parse_time(minutes, seconds):
    """Convert to float seconds."""
    t = float(seconds)
    if minutes:
        t += int(minutes) * 60
    return round(t, 3)


def find_cup_id(ws, row, col):
    """Look backward from cell to find the cup header (row 2 typically)."""
    for cc in range(col, max(0, col - 6), -1):
        for rr in (2, 3, 1):  # row 2 most common, but check nearby
            val = ws.cell(rr, cc).value
            if val and isinstance(val, str):
                val = val.strip()
                if any(k in val for k in ('COTD', 'COTW', 'Roulette', 'Troll')):
                    return val
    return None


def find_max_round(ws, ft_row, ft_col):
    """Find the max Elim Round value in the same cup section as the FT cell.
    Elim Round is typically in column ft_col+1 (offset from Position col).
    We search the Position column area and find max elim round."""
    # Find the Position header row (usually row 5, near ft_row+1)
    pos_col = None
    header_row = None
    for rr in range(ft_row, ft_row + 3):
        for cc in range(max(1, ft_col - 4), ft_col + 2):
            v = ws.cell(rr, cc).value
            if v and str(v).strip() == 'Position':
                pos_col = cc
                header_row = rr
                break
        if pos_col:
            break
    if not pos_col:
        return None

    # Elim Round column is typically pos_col + 3
    eround_col = pos_col + 3
    max_r = 0
    for rr in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(rr, pos_col + 1).value
        if name is None:
            break
        er = ws.cell(rr, eround_col).value
        if er and isinstance(er, (int, float)):
            max_r = max(max_r, int(er))
    return max_r if max_r > 0 else None


def scan_file(filepath):
    """Scan a single xlsx for all Fastest Time entries."""
    results = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  SKIP {os.path.basename(filepath)}: {e}")
        return results

    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if not val or not isinstance(val, str) or 'Fastest Time:' not in val:
                    continue
                # Skip irrelevant notes
                if 'irrelevent' in val.lower() or 'irrelevant' in val.lower():
                    continue

                m = RE_FT.match(val)
                if not m:
                    continue

                minutes, seconds, player, round_info = m.groups()
                time_val = parse_time(minutes, seconds)
                player = player.strip()
                round_info = round_info.strip() if round_info else None
                cup_id = find_cup_id(ws, r, c)

                if not cup_id:
                    cup_id = sname  # fallback to sheet name

                # Compute players remaining when FT was set
                max_round = find_max_round(ws, r, c)
                players_left = None
                if round_info == 'Final':
                    players_left = 2
                elif max_round and round_info:
                    rm = re.search(r'(\d+)', round_info)
                    if rm:
                        ft_round = int(rm.group(1))
                        players_left = max_round - ft_round + 2

                entry = {
                    'cup': cup_id,
                    'time': time_val,
                    'player': normalize_name(player),
                    'round': round_info,
                }
                if players_left:
                    entry['left'] = players_left

                results.append(entry)
    wb.close()
    return results


# Scan all files
all_entries = []
for fname in FILES:
    path = _p(fname)
    if not os.path.exists(path):
        print(f"  WARNING: {fname} not found, skipping")
        continue
    entries = scan_file(path)
    print(f"  {fname}: {len(entries)} fastest times")
    all_entries.extend(entries)

print(f"\nTotal raw entries: {len(all_entries)}")

# Deduplicate by cup ID — prefer entries with round info
by_cup = {}
for e in all_entries:
    cup = e['cup']
    if cup not in by_cup:
        by_cup[cup] = e
    else:
        existing = by_cup[cup]
        # Prefer entry with round info and players_left data
        has_more = (e.get('round') and not existing.get('round')) or \
                   (e.get('left') and not existing.get('left'))
        if has_more:
            by_cup[cup] = e
        elif e.get('round') and existing.get('round') and e['time'] < existing['time']:
            by_cup[cup] = e

fastest = sorted(by_cup.values(), key=lambda x: x['time'])

print(f"After dedup: {len(fastest)} cups with fastest times")

# Write output
with open(_p('fastest.json'), 'w', encoding='utf-8') as f:
    json.dump(fastest, f, separators=(',', ':'), ensure_ascii=False)

print(f"Wrote fastest.json ({len(fastest)} entries)")

# Show top 10
print("\nTop 10 fastest:")
for i, e in enumerate(fastest[:10], 1):
    rd = e['round'] or '?'
    print(f"  {i}. {e['time']:.3f}s  {e['player']:<25} {rd:<15} {e['cup']}")
