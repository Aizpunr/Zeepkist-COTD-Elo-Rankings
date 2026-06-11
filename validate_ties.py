"""Validate that DNFs in the same round share the same position."""
import openpyxl, sys, os, re
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

_dir = os.path.dirname(os.path.abspath(__file__))
_p = lambda f: os.path.join(_dir, f)

# Read the xlsx list from elo_engine.py source — never hardcode it here.
# A hardcoded list silently went stale after the COTD 131-138 -> 131-139
# rename and this script validated nothing past cup 130 for months.
_elo_src = open(_p('elo_engine.py'), encoding='utf-8').read()
files = re.findall(r"parse_file\(_p\('(.+?\.xlsx)'\)\)", _elo_src)
files += re.findall(r"parse_troll_cups\(_p\('(.+?\.xlsx)'\)\)", _elo_src)
if not files:
    sys.exit("ERROR: no xlsx references found in elo_engine.py")

issues = 0
cups_checked = 0

for fname in files:
    try:
        wb = openpyxl.load_workbook(_p(fname), data_only=True)
    except FileNotFoundError:
        sys.exit(f"ERROR: {fname} is referenced by elo_engine.py but missing on disk")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if not val or not str(val).startswith(('COTD', 'COTW')):
                    continue

                # Find Position header
                pos_row = None
                for r2 in range(ri, min(ri + 5, len(rows))):
                    if ci < len(rows[r2]) and rows[r2][ci] == 'Position':
                        pos_row = r2
                        break
                if not pos_row:
                    continue

                # Check if this cup has Time + Round columns
                has_time = (ci + 2 < len(rows[pos_row])
                            and rows[pos_row][ci + 2] in ('Elim Time', 'Time'))
                has_round = (ci + 3 < len(rows[pos_row])
                             and rows[pos_row][ci + 3] in ('Elim Round', 'Round'))

                # Read players
                players = []
                last_pos = None
                for row2 in rows[pos_row + 1:]:
                    if ci >= len(row2) or ci + 1 >= len(row2):
                        break
                    pos = row2[ci]
                    name = row2[ci + 1]
                    if name is None:
                        break

                    if pos is not None:
                        try:
                            last_pos = int(float(str(pos).rstrip('*').strip()))
                        except ValueError:
                            continue
                    p = last_pos

                    time_val = row2[ci + 2] if has_time and ci + 2 < len(row2) else None
                    rnd_val = row2[ci + 3] if has_round and ci + 3 < len(row2) else None

                    players.append({
                        'pos': p,
                        'name': str(name).strip(),
                        'time': str(time_val).strip() if time_val else None,
                        'round': int(float(str(rnd_val))) if rnd_val else None,
                    })

                cups_checked += 1
                cup_name = str(val).strip()

                # --- Check 1: same round + DNF → must share position ---
                if has_time and has_round:
                    by_round = defaultdict(list)
                    for p in players:
                        if p['round'] and p['time'] == 'DNF':
                            by_round[p['round']].append(p)

                    for rnd, group in by_round.items():
                        positions = set(p['pos'] for p in group)
                        if len(positions) > 1:
                            issues += 1
                            names = ', '.join(p['name'] for p in group)
                            print(f"[DNF] {cup_name} round {rnd}: "
                                  f"{len(group)} DNFs at positions "
                                  f"{sorted(positions)} — {names}")

                # NOTE: there used to be a Check 2 here ("everyone eliminated in
                # the same round must share a position"). That was the overzealous
                # 2026-04-07 tie rule, corrected 2026-04-12: finishers in a round
                # get DISTINCT positions ordered by elim time; only DNFs tie.
                # Check 1 above is the only check valid under the current rule.

print(f"\n{cups_checked} cups checked, {issues} issues found")
